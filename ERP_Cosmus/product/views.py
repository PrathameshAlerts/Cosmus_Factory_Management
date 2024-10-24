import datetime
import decimal
from io import BytesIO
from operator import itemgetter
import os
from re import S
from django.conf import settings

from django.core.exceptions import ValidationError , ObjectDoesNotExist
import json
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q, Sum, ProtectedError, Avg, Count,F,Exists, OuterRef, Subquery, DecimalField
from django.db.models.functions import Round
from django.db import DatabaseError, IntegrityError, transaction
from django.contrib.auth.decorators import login_required, user_passes_test
import logging
import urllib.parse
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.forms import inlineformset_factory, modelformset_factory
from django.http import HttpResponse, HttpResponseServerError, JsonResponse
from django.views.decorators.cache import cache_control
from django.db.models.functions import Coalesce
import pandas as pd
from openpyxl.drawing.image import Image  
from PIL import Image as PILImage
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side


from . models import (AccountGroup, AccountSubGroup, Color, Fabric_Group_Model,
                       FabricFinishes, Godown_finished_goods, Godown_raw_material,
                         Item_Creation, Ledger, MainCategory, PProduct_Creation, Product,
                           Product2SubCategory,  ProductImage, RawStockTransferMaster, StockItem,
                             SubCategory, Unit_Name_Create, account_credit_debit_master_table, cutting_room, factory_employee, godown_item_report_for_cutting_room,
                               gst, item_color_shade, item_godown_quantity_through_table,
                                 item_purchase_voucher_master, labour_work_in_master, labour_work_in_product_to_item, labour_workout_childs, labour_workout_cutting_items, labour_workout_master, ledgerTypes, opening_shade_godown_quantity, 
                                 packaging, product_2_item_through_table, product_godown_quantity_through_table, product_to_item_labour_child_workout, product_to_item_labour_workout, purchase_order, 
                                 purchase_order_for_raw_material, purchase_order_raw_material_cutting, 
                                 purchase_order_to_product, purchase_order_to_product_cutting, purchase_voucher_items,
                                   set_prod_item_part_name, shade_godown_items,
                                   shade_godown_items_temporary_table,purchase_order_for_raw_material_cutting_items)

from .forms import( Basepurchase_order_for_raw_material_cutting_items_form, ColorForm, 
                    CustomPProductaddFormSet, ProductCreateSkuFormsetCreate,
                     ProductCreateSkuFormsetUpdate, cutting_room_form,
                       factory_employee_form, labour_work_in_product_to_item_approval_formset, labour_work_in_product_to_item_form, labour_workin_master_form, labour_workout_child_form, labour_workout_cutting_items_form, ledger_types_form, purchase_order_for_raw_material_cutting_items_form, 
                       purchase_order_to_product_cutting_form,raw_material_stock_trasfer_items_formset,
                    FabricFinishes_form, ItemFabricGroup, Itemform, LedgerForm,
                     OpeningShadeFormSetupdate, PProductAddForm, PProductCreateForm, ShadeFormSet,
                       StockItemForm, UnitName, account_sub_grp_form, PProductaddFormSet,
                        ProductImagesFormSet, ProductVideoFormSet, purchase_order_form,purchase_voucher_items_godown_formset_shade_change,
                         gst_form, item_purchase_voucher_master_form,
                           packaging_form, product_main_category_form,  Product2ItemFormsetExtraForm,Product2CommonItemFormSetExtraForm,
                            product_sub_category_form, purchase_voucher_items_formset,
                             purchase_voucher_items_godown_formset, purchase_voucher_items_formset_update, raw_material_stock_trasfer_master_form,
                                shade_godown_items_temporary_table_formset,shade_godown_items_temporary_table_formset_update,
                                Product2ItemFormset,Product2CommonItemFormSet,purchase_order_product_qty_formset,
                                purchase_order_raw_product_qty_formset,purchase_order_raw_product_qty_cutting_formset,
                                purchase_order_cutting_approval_formset,
                                purchase_order_raw_product_sheet_form,purchase_order_raw_material_cutting_form)


logger = logging.getLogger('product_views')


"""
    logger.debug("This is a debug message")
    logger.info("This is an info message")
    logger.warning("This is a warning message")
    logger.error("This is an error message")
    logger.critical("This is a critical message")
"""

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)




@login_required(login_url='login')
def dashboard(request):
    return render(request,'misc/dashboard.html')




@login_required(login_url='login')
def edit_production_product(request,pk):
   
    gsts = gst.objects.all()
    pproduct = get_object_or_404(Product, Product_Refrence_ID=pk)
    
    
    product_skus = pproduct.productdetails.all()
    
    products_sku_counts = PProduct_Creation.objects.filter(Product__Product_Refrence_ID=pk).count()

    
    prod2cat_instance = Product2SubCategory.objects.filter(Product_id= pproduct.id)
    prod_main_cat_name = ''
    prod_main_cat_id = ''
    prod_sub_cat_dict = {}
    prod_sub_cat_dict_all = {}

    
    if prod2cat_instance.exists():
        prodmaincat = prod2cat_instance.first()
        
        prod_main_cat_name = prodmaincat.SubCategory_id.product_main_category.product_category_name
        prod_main_cat_id = prodmaincat.SubCategory_id.product_main_category.id


        
        for subcat in prod2cat_instance:
            prod_sub_cat_dict[subcat.SubCategory_id.id] = subcat.SubCategory_id.product_sub_category_name


        
        sub_categories = SubCategory.objects.filter(product_main_category = prod_main_cat_id)
        for sub_cat_all in sub_categories:
            prod_sub_cat_dict_all[sub_cat_all.id] = sub_cat_all.product_sub_category_name


    colors = Color.objects.all()
    main_categories = MainCategory.objects.all()

    form = PProductAddForm(instance=pproduct)
    formset = CustomPProductaddFormSet(instance=pproduct)

    if request.method == 'POST':
        product_ref_id = pk
        
        try:
            excel_file = request.FILES.get('excel_file')

            if excel_file:
                file_name = excel_file.name
                product_ref_id_file = file_name.split('_')[-1].split('.')[0].split(' ')[0]
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, 'Invalid file format. Please upload a valid Excel file.')
                    logger.error('Invalid file format. Please upload a valid Excel file.')
                    return redirect('pproductlist')
                
                if int(product_ref_id_file) == product_ref_id:
                    with transaction.atomic():
                        wb = load_workbook(excel_file, data_only = True)
                        ws1 = wb['product_special_configs']
                        ws2 = wb['product_common_configs']

                        
                        grand_total = 0
                        grand_total_combi = 0

                        for row in ws1.iter_rows(min_row=2,min_col=1):
                            id = row[0].value
                            item_name = row[1].value
                            product_sku = row[2].value
                            body_combi = row[7].value
                           
                            if id is not None:
                                if product_sku is not None and item_name is not None:
                                    part_name = row[3].value
                                    part_dimention = row[4].value
                                    dimention_total = row[5].value
                                    part_pieces = row[6].value
                                    body_combi = row[7].value
                                    
                                    if part_name is not None and part_dimention is not None and body_combi is not None:   

                                        if body_combi == 'body':
                                            grand_total = grand_total + float(dimention_total)  

                                        elif body_combi == 'combi':
                                            grand_total_combi = grand_total_combi + float(dimention_total)  


                                        p2i_config_instance = set_prod_item_part_name.objects.get(id=id)

                                        p2i_config_instance.producttoitem.grand_total = grand_total   
                                        p2i_config_instance.producttoitem.grand_total_combi = grand_total_combi   
                                        p2i_config_instance.part_name = part_name
                                        p2i_config_instance.part_dimentions = part_dimention
                                        p2i_config_instance.dimention_total = dimention_total
                                        p2i_config_instance.part_pieces = part_pieces
                                        p2i_config_instance.body_combi = body_combi
                                        p2i_config_instance.c_user = request.user
                                        p2i_config_instance.producttoitem.c_user = request.user
                                        p2i_config_instance.save()   
                                        p2i_config_instance.producttoitem.save()  

                                    else:
                                        p2i_config_instance = set_prod_item_part_name.objects.get(id = id)  
                                        p2i_config_instance.producttoitem.no_of_rows = p2i_config_instance.producttoitem.no_of_rows - 1   
                                        p2i_config_instance.producttoitem.c_user =  request.user
                                        p2i_config_instance.delete()
                                        p2i_config_instance.producttoitem.save()


                            else:
                                grand_total = 0
                                grand_total_combi = 0

                        
                        for product_c in PProduct_Creation.objects.filter(Product__Product_Refrence_ID = pk): 
                            product_sku = product_c.PProduct_SKU
                           
                            row_no = 0  
                            grand_total = 0
                            grand_total_combi = 0

                            for row in ws2.iter_rows(min_row=2,min_col=1):  
                                
                                id = row[0].value
                                item_name = row[1].value
                                part_name = row[2].value
                                part_dimention = row[3].value
                                dimention_total = row[4].value
                                part_pieces = row[5].value
                                body_combi = row[6].value
                                
                                if id is not None and item_name is not None:  
                                    
                                    
                                    p2i_instances = product_2_item_through_table.objects.get(PProduct_pk = product_sku, Item_pk__item_name = item_name, common_unique = True)
                                    
                                    
                                    
                                    
                                    
                                    p2i_instances_configs = set_prod_item_part_name.objects.filter(producttoitem=p2i_instances).order_by('id')[row_no]
                                    

                                    if part_name is not None and body_combi is not None:  

                                        if body_combi == 'body':
                                            grand_total = grand_total + float(dimention_total)   
                                    
                                        elif body_combi == 'combi':
                                            grand_total_combi = grand_total_combi + float(dimention_total)  

                                        p2i_instances_configs.part_name = part_name
                                        p2i_instances_configs.part_dimentions = part_dimention
                                        p2i_instances_configs.part_pieces = part_pieces
                                        p2i_instances_configs.body_combi = body_combi
                                        p2i_instances_configs.dimention_total = dimention_total
                                        p2i_instances_configs.producttoitem.grand_total = grand_total 
                                        p2i_instances_configs.producttoitem.grand_total_combi = grand_total_combi   


                                        p2i_instances_configs.c_user = request.user
                                        p2i_instances_configs.producttoitem.c_user = request.user
                                        p2i_instances_configs.save()
                                        p2i_instances_configs.producttoitem.save() 
                                        row_no = row_no + 1  

                                    else:

                                        p2i_instances_configs.producttoitem.no_of_rows = p2i_instances_configs.producttoitem.no_of_rows - 1
                                        p2i_instances_configs.producttoitem.c_user = request.user
                                        p2i_instances_configs.delete()
                                        p2i_instances_configs.producttoitem.save()
                                        

                                else:
                                    row_no = 0
                                    grand_total = 0
                                    grand_total_combi = 0


                else:
                    messages.error(request, 'File with invalid Product Refrence Id uploaded')
                    logger.error("File with invalid Product Refrence Id uploaded")
                    return redirect('pproductlist')
            
        except Exception as e:
            logger.error(f"An error occured - {str(e)} - Product-Name - {pproduct}")
            messages.error(request, f'Error uploading Excel file: {str(e)}')
            return redirect('pproductlist')
        
        form = PProductAddForm(request.POST, request.FILES, instance = pproduct) 
        formset = CustomPProductaddFormSet(request.POST, request.FILES , instance=pproduct)
        
        formset.forms = [form for form in formset.forms if form.has_changed()]
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    form_instance = form.save(commit=False)
                    form_instance.c_user = request.user
                    
                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()

                    for form in formset:
                        if form.is_valid():
                            if not form.cleaned_data.get('DELETE'):
                                formset_instances = form.save(commit=False)
                                formset_instances.c_user = request.user
                                formset_instances.save()

                    
                    p_id_pk = form_instance.pk
                    p_id = Product.objects.get(pk=p_id_pk)
        
                    
                    sub_category_ids = request.POST.getlist('Product_Sub_catagory')
            
                    
                    sub_cat_front_listcomp = [Product2SubCategory.objects.filter(Product_id=p_id,SubCategory_id=sub_cat_id).first() for sub_cat_id in sub_category_ids]
            
                    
                    sub_cat_backend = [x for x in Product2SubCategory.objects.filter(Product_id=p_id)]

                    
                    objects_to_delete = [obj for obj in sub_cat_backend if obj not in sub_cat_front_listcomp]
            
                    for obj in objects_to_delete:
                        obj.delete()

                    
                    for sub_cat_id in sub_category_ids:
                        sub_cat = SubCategory.objects.get(id = sub_cat_id)
                        p2c, created = Product2SubCategory.objects.get_or_create(Product_id=p_id, SubCategory_id=sub_cat)
                        p2c.c_user = request.user
                        p2c.save()

                    
                    form_instance.save()
                    logger.info(f"product-saved product")
                    logger.info(f"product-formset-saved product")
                    return redirect('pproductlist')

            except Exception as e:
                logger.error(f"An exception occured in Product - {e} - Product-name - {pproduct}")
                messages.error(request, f'An exception occured - {e}')
        
        else:

            logger.error(f"Productform not valid - {form.errors} - Product-name - {pproduct}")
            logger.error(f"Product formsets not valid- {formset.errors} - Product-name - {pproduct}")

            return render(request, 'product/edit_production_product.html', {'gsts':gsts,
                                                                            'form':form,
                                                                            'product_skus':product_skus,
                                                                            'formset':formset,
                                                                            'colors':colors,
                                                                            'products_sku_counts':products_sku_counts,
                                                                            'main_categories':main_categories,
                                                                            'prod_main_cat_name':prod_main_cat_name,
                                                                            'prod_main_cat_id':prod_main_cat_id,
                                                                            'prod_sub_cat_dict':prod_sub_cat_dict,
                                                                            'prod_sub_cat_dict_all':prod_sub_cat_dict_all})


    return render(request, 'product/edit_production_product.html',{'gsts':gsts,
                                                                   'form': form,
                                                                   'product_skus':product_skus,
                                                                   'formset':formset,
                                                                   'colors':colors,
                                                                   'products_sku_counts':products_sku_counts,
                                                                   'main_categories':main_categories,
                                                                   'prod_main_cat_name':prod_main_cat_name,
                                                                    'prod_main_cat_id':prod_main_cat_id,
                                                                    'prod_sub_cat_dict':prod_sub_cat_dict,
                                                                    'prod_sub_cat_dict_all':prod_sub_cat_dict_all})




@login_required(login_url='login')
def product2subcategoryproductajax(request):
    selected_main_cat = request.GET.get('p_main_cat')
    sub_cats = SubCategory.objects.filter(product_main_category = selected_main_cat)
    
    sub_cat_dict = {}

    for sub_cat in sub_cats:
        sub_cat_dict[sub_cat.id] = sub_cat.product_sub_category_name 


    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse({'sub_cat_dict':sub_cat_dict})





@login_required(login_url='login')
def product_color_sku(request,ref_id = None):

    color = Color.objects.all()

    if ref_id:
        instance = Product.objects.get(Product_Refrence_ID=ref_id) 
        
        ordered_product_details = instance.productdetails.all().order_by('created_date')
        
        formset = ProductCreateSkuFormsetUpdate(request.POST or None, request.FILES or None, instance=instance, c_user=request.user)
        
    else:
        instance = None
        formset = ProductCreateSkuFormsetCreate(request.POST or None, request.FILES or None, instance=instance, c_user=request.user)

    

    if request.method == 'POST':
        product_ref_id = request.POST.get('Product_Refrence_ID', None)
        
        
        formset.forms = [form for form in formset if form.has_changed()]

        if product_ref_id:
            if formset.is_valid():
                try:
                    for form in formset:
                        if form.is_valid():
                            form_instance = form.save(commit=False)  
                            obj, created =  Product.objects.get_or_create(Product_Refrence_ID=product_ref_id) 
                            obj.c_user = request.user
                            obj.save()
                            form_instance.Product = obj   
                            form_instance.save() 

                    return redirect(reverse('edit_production_product', args=[product_ref_id]))
                        
                except ValidationError as ve :
                    messages.error(request,f'{ve}')
                    logger.error(ve)

                except Exception as e :
                    messages.error(request,f'{e}')
                    logger.error(e)
            else:
                return render(request, 'product/product_color_sku.html', {'formset': formset, 'color': color,'ref_id': ref_id})


    return render(request, 'product/product_color_sku.html', {'formset': formset, 'color': color,'ref_id': ref_id})




@login_required(login_url='login')
def pproduct_list(request):
    
    queryset = Product.objects.all().order_by('Product_Name').select_related('Product_GST').prefetch_related('productdetails','productdetails__PProduct_color')

    product_search = request.GET.get('product_search','')
    
    if product_search != '':
        queryset = Product.objects.filter(Q(Product_Name__icontains=product_search)|
                                            Q(Model_Name__icontains=product_search)|
                                            Q(Product_Refrence_ID__icontains=product_search)|
                                            Q(productdetails__PProduct_SKU__icontains=product_search)).distinct()
    
    
    paginator = Paginator(queryset, 10)  
    
    
    page_number = request.GET.get('page')
    
    try:
        products = paginator.page(page_number)
    except PageNotAnInteger:
        
        products = paginator.page(1)
    except EmptyPage:
        
        products = paginator.page(paginator.num_pages)

    
    index = products.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 2 if index >= 2 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_range = list(paginator.page_range)[start_index:end_index]

    context = {
        'products': products,
        'page_range': page_range,
        'product_search':product_search
    }

    return render(request,'product/pproduct_list.html',context=context)


@login_required(login_url='login')
def pproduct_delete(request, pk):
    try:
        product = get_object_or_404(Product,Product_Refrence_ID=pk)
        product.delete()
        messages.success(request,f'Product {product.Product_Name} was deleted')
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {product.Product_Name} because it is referenced by other objects.')
    return redirect('pproductlist')



@login_required(login_url='login')
def add_product_images(request, pk):
    product = PProduct_Creation.objects.get(pk=pk)   
    formset = ProductImagesFormSet(instance=product)  
    
    if request.method == 'POST':
        formset = ProductImagesFormSet(request.POST, request.FILES, instance=product, c_user=request.user)
        if formset.is_valid():
            formset.save()
            messages.success(request,'Product images sucessfully added.')
            

            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """

            return HttpResponse(close_window_script)
        else:
            return render(request, 'product/add_product_images.html', {'formset': formset, 'product': product})

    return render(request, 'product/add_product_images.html', {'formset': formset, 'product': product})


@login_required(login_url='login')
def add_product_video_url(request,pk):
    
    product = PProduct_Creation.objects.get(pk=pk)   
    formset = ProductVideoFormSet(instance= product)  

    if request.method == 'POST':
        formset = ProductVideoFormSet(request.POST, instance=product, c_user=request.user)
        
        if formset.is_valid():
            formset.save()
            messages.success(request,'Product url sucessfully added.')
            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """
            return HttpResponse(close_window_script)

    else:
            return render(request, 'product/add_product_videourl.html', {'formset': formset, 'product': product})

    return render(request, 'product/add_product_videourl.html', {'formset': formset, 'product': product})


@login_required(login_url='login')
def definemaincategoryproduct(request,pk=None):

    queryset = MainCategory.objects.all()

    main_cat_product_search = request.GET.get('main_cat_product_search','')

    if main_cat_product_search != '':
        queryset = MainCategory.objects.filter(product_category_name__icontains = main_cat_product_search)


    if pk:
        instance = MainCategory.objects.get(pk=pk)
        title = 'Update'
        message = 'updated'
    else:
        instance = None
        title = 'Create'
        message = 'created'

    
    form = product_main_category_form(instance=instance)
    if request.method == 'POST':
        form = product_main_category_form(request.POST, instance= instance)
        if form.is_valid():
            form_instance  = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()
            if message == 'created':
                messages.success(request,'Main Category created sucessfully')
            if message == 'updated':
                messages.success(request,'Main Category updated sucessfully')
            return redirect('define-main-category-product')
        else:
            return render(request,'product/definemaincategoryproduct.html',{'form':form,'main_cats':queryset,
                                                                    'title':title,'main_cat_product_search':main_cat_product_search})
        
    return render(request,'product/definemaincategoryproduct.html',{'form':form,'main_cats':queryset,
                                                                    'title':title,'main_cat_product_search':main_cat_product_search})


@login_required(login_url='login')
def definemaincategoryproductdelete(request,pk):
    try:
        instance = MainCategory.objects.get(pk=pk)
        instance.delete()
        messages.success(request,'Main Category Deleted Successfully.')
    except Exception as e :
        messages.error(request,f'{e}')
    return redirect('define-main-category-product')



@login_required(login_url='login')
def definesubcategoryproduct(request, pk=None):
    if pk:
        instance = SubCategory.objects.get(pk=pk)
        title = 'Update'
        message = 'updated'
    else:
        instance = None
        title = 'Create'
        message = 'created'

    main_categories = MainCategory.objects.all()
    sub_category = SubCategory.objects.all()
    
    
    form = product_sub_category_form(instance = instance)
    if request.method == 'POST':
        
        try:
            form = product_sub_category_form(request.POST,instance = instance)
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.c_user = request.user
                form_instance.save()
                if message == 'created':
                    messages.success(request,'Sub-Category created sucessfully')
                if message == 'updated':
                    messages.success(request,'Sub-Category updated sucessfully')
            
                return redirect('define-sub-category-product')
            else:
                print(form.errors)
                messages.error(request,f'An Exception occoured - {form.errors}')
        
        except Exception as e:
            messages.error(request,f'An Exception occoured - {e}')

    return render(request,'product/definesubcategoryproduct.html',{'main_categories':main_categories, 'sub_category':sub_category,'form':form,'title':title})


@login_required(login_url='login')
def definesubcategoryproductdelete(request, pk):
    try:
        instance = SubCategory.objects.get(pk=pk)
        instance.delete()
        messages.success(request,'Sub Category Deleted Successfully.')
    except Exception as e:
        messages.error(request,f'An Exception occoured - {e}')    
    return redirect('define-sub-category-product')



@login_required(login_url='login')
def product2subcategory(request):
    products = Product.objects.all()
    sub_category = SubCategory.objects.all()
    main_categories = MainCategory.objects.all()
    
    if request.method == 'POST':

        try:
            
            product_id_get = request.POST.get('product_name')
            
            
            sub_category_ids = request.POST.getlist('sub_category_name')
            
            
            p_id = get_object_or_404(Product, id = product_id_get)

            
            existing_instances =  Product2SubCategory.objects.filter(Product_id=p_id)

            updated_instances_front = []
            
            
            for sub_cat_id in sub_category_ids:
                
                s_c_id =  get_object_or_404(SubCategory, id = sub_cat_id)
                
                p_2_c_instance = Product2SubCategory.objects.filter(Product_id=p_id, SubCategory_id=s_c_id).first() 
                updated_instances_front.append(p_2_c_instance)

            
            
            updated_instance_pk = set(obj.pk for obj in updated_instances_front if obj is not None)
            
            instances_to_delete = [obj for obj in existing_instances if obj.pk not in updated_instance_pk]
            
            
            for obj in instances_to_delete:
                obj.delete()

            
                
            
            
            for sub_cat_id in sub_category_ids:

                
                s_c_id =  get_object_or_404(SubCategory, id = sub_cat_id)

                p2c, created = Product2SubCategory.objects.get_or_create(Product_id=p_id, SubCategory_id=s_c_id)
                p2c.c_user = request.user
                p2c.save()
            messages.success(request,f'Product sucessfully added to {s_c_id.product_sub_category_name}')
        
        except IntegrityError:
            messages.error(request, 'Product already present in Subcategory')

        except Exception as e:
            messages.error(request,f'An Exception occoured - {e}')

    return render(request,'product/product2subcategory.html',{'main_categories':main_categories,'products':products,'sub_category':sub_category})



@login_required(login_url='login')
def product2subcategoryajax(request):

    productid = request.GET.get('selected_product_id')
    categoryforselectedproduct = Product2SubCategory.objects.filter(Product_id = productid)
    
    dict_result = {}

    for obj in categoryforselectedproduct:
        dict_result[obj.SubCategory_id.id] = obj.SubCategory_id.product_sub_category_name
    
    

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({'dict_result':dict_result})
    







@login_required(login_url='login')
def item_create(request):

    title = 'Item Create'
    gsts = gst.objects.all()
    fab_grp = Fabric_Group_Model.objects.all()
    unit_name = Unit_Name_Create.objects.all()
    packaging_material_all = packaging.objects.all()
    fab_finishes = FabricFinishes.objects.all()
    items_to_clone = Item_Creation.objects.all()
    colors = Color.objects.all()
    form = Itemform()
    

    if request.path == '/itemcreatepopup/':
        template_name = 'product/item_create_popup.html'

    else:
        template_name = 'product/item_create_update.html'
    
    if request.method == 'POST':
        form = Itemform(request.POST, request.FILES)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()
            if request.path == '/itemcreatepopup/':
                return HttpResponse('item created', status = 200) 
            else:
                logger.info("Item Successfully Created")
                messages.success(request,'Item has been created')
                return redirect(reverse('item-edit', args=[form_instance.id]))
    
        else:
            logger.error(f"item form not valid{form.errors}")
            messages.error(request, f"item form not valid{form.errors}")
           
            return render(request,template_name, {'gsts':gsts,
                                                                      'fab_grp':fab_grp,
                                                                      'unit_name':unit_name,
                                                                      'colors':colors,
                                                                      'packaging_material_all':packaging_material_all,
                                                                      'fab_finishes':fab_finishes,
                                                                      'title':title,'form':form,'items_to_clone':items_to_clone})
    
    
    return render(request,template_name,{'gsts':gsts,
                                                                 'fab_grp':fab_grp,
                                                                 'unit_name':unit_name,
                                                                 'colors':colors,
                                                                 'title':title,
                                                                 'packaging_material_all':packaging_material_all,
                                                                    'fab_finishes':fab_finishes,
                                                                 'form':form,'items_to_clone':items_to_clone})






    

@login_required(login_url='login')
def item_clone_ajax(request):
    selected_item_name_value = int(request.GET.get('itemValue'))
    
    if selected_item_name_value:
        selected_item = get_object_or_404(Item_Creation, pk=selected_item_name_value)

        response_data = {'fabric_group':{'fab_g_key':selected_item.Fabric_Group.id,'fab_g_value':selected_item.Fabric_Group.fab_grp_name},
                         'color':{'color_key':selected_item.Item_Color.id,'color_value':selected_item.Item_Color.color_name}, 
                         'material_code':selected_item.Material_code,'packing':{'packing_key':selected_item.Item_Packing.id ,'packing_value':selected_item.Item_Packing.packing_material},
                        'unit_name':{'unit_name_key':selected_item.unit_name_item.id,'unit_name_value':selected_item.unit_name_item.unit_name},
                        'panha':selected_item.Panha,'fab_non_fab':selected_item.Fabric_nonfabric,
                        'fab_finishes':{'fab_finishes_key':selected_item.Item_Fabric_Finishes.id,'fab_finishes_value':selected_item.Item_Fabric_Finishes.fabric_finish},
                        'gst':{'gst_key':selected_item.Item_Creation_GST.id,'gst_value':selected_item.Item_Creation_GST.gst_percentage},
                        'hsn_code':selected_item.HSN_Code,'status':selected_item.status, 'unit_name_units':selected_item.unit_name_item.unit_value}
       
    return JsonResponse({'response_data':response_data})




@login_required(login_url='login')
def item_list(request):
    
    g_search = request.GET.get('item_search','')

    
    
    queryset = Item_Creation.objects.all().annotate(total_quantity=Sum('shades__godown_shades__quantity')).order_by('item_name').select_related('Item_Color','unit_name_item',
                                                    'Fabric_Group','Item_Creation_GST','Item_Fabric_Finishes',
                                                    'Item_Packing').prefetch_related('shades',
                                                    'shades__godown_shades')

    
    if g_search != '':
        queryset = queryset.filter(Q(item_name__icontains=g_search)|
                                                Q(Item_Color__color_name__icontains=g_search)|
                                                Q(Fabric_Group__fab_grp_name__icontains=g_search))
        

    sort_name = request.GET.get('sort_name')


    if sort_name == "item_name_sort_asc" :
        queryset = Item_Creation.objects.order_by('item_name')
    

    elif sort_name == "item_name_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-item_name')


    elif sort_name == "fabgrp_sort_asc" :
        queryset = Item_Creation.objects.order_by('Fabric_Group__fab_grp_name')


    elif sort_name == "fabgrp_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-Fabric_Group__fab_grp_name')

    elif sort_name == "item_color_sort_dsc" :
        queryset = Item_Creation.objects.order_by('Item_Color__color_name')
    
    elif sort_name == "item_color_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-Item_Color__color_name')

    any_desc = request.GET.get('any_desc')
    exact_desc = request.GET.get('exact_desc')

    if any_desc:
        if any_desc != '' and any_desc is not None:
            queryset = Item_Creation.objects.filter(item_name__icontains=any_desc)
        
    if exact_desc:
        if exact_desc != '' and exact_desc is not None:
            queryset = Item_Creation.objects.filter(item_name__exact=exact_desc)

    return render(request,'product/list_item.html', {"items":queryset,"item_search":g_search})
    



@login_required(login_url='login')
def item_edit(request,pk): 
    
    title = 'Item update'
    gsts = gst.objects.all()
    fab_grp = Fabric_Group_Model.objects.all()
    unit_name = Unit_Name_Create.objects.all()
    colors = Color.objects.all()
    packaging_material_all = packaging.objects.all()
    fab_finishes = FabricFinishes.objects.all()
    item_pk = get_object_or_404(Item_Creation ,pk = pk)

    form = Itemform(instance=item_pk)

    
    queryset = item_color_shade.objects.filter(items = pk).annotate(total_quantity=Sum('opening_shade_godown_quantity__opening_quantity'),
                                                                     total_value=Sum(F('opening_shade_godown_quantity__opening_quantity') * F('opening_shade_godown_quantity__opening_rate'), 
                                                                                output_field=DecimalField(max_digits=10, decimal_places=2)))

    formset = ShadeFormSet(instance= item_pk, queryset=queryset)

    
    
    if request.method == 'POST':
        form = Itemform(request.POST, request.FILES , instance = item_pk)
        formset = ShadeFormSet(request.POST , request.FILES, instance = item_pk)
        formset.forms = [form for form in formset if form.has_changed()] 
        try:
            if form.is_valid() and formset.is_valid():
                form_instance = form.save(commit=False)  
                form_instance.c_user = request.user
                form_instance.save()


                for form in formset.deleted_forms: 
                    if form.instance.pk:
                        form.instance.delete()

                for form in formset: 
                    if form.is_valid():
                        if form.instance.pk: 
                            form_instance = form.save(commit=False)
                            form_instance.c_user = request.user
                            form_instance.save()

                        else:  
                            
                            if not form.cleaned_data.get('DELETE'):
                                shade_form_instance = form.save(commit=False) 
                                shade_form_instance.c_user = request.user
                                shade_form_instance.save() 

                                form_prefix_number = form.prefix[-1] 
                                opening_godown_quantity = request.POST.get(f'shades-{form_prefix_number}-openingValue') 

                                if opening_godown_quantity != '':
                                    opening_godown_quantity_dict = json.loads(opening_godown_quantity)
                                    opening_godown_qty_data = opening_godown_quantity_dict.get('newData')
                                    
                                    
                                    item_godown_formset_data = {}
                                    for key , value in opening_godown_qty_data.items():
                                        form_set_id = key.split('_')[-1]

                                        new_data_get = opening_godown_qty_data.get(key)
                                        
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-TOTAL_FORMS'] = str(len(opening_godown_qty_data))
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-INITIAL_FORMS'] =  str(0)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-MIN_NUM_FORMS'] =  str(0)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-MAX_NUM_FORMS'] =  str(1000)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_godown_id'] = new_data_get.get('godownData')
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_quantity'] = new_data_get.get('qtyData')
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_rate'] = new_data_get.get('rateData')
                                    
                                    
                                    new_godown_opening_formsets = OpeningShadeFormSetupdate(item_godown_formset_data, prefix='opening_shade_godown_quantity_set')

                                    for form in new_godown_opening_formsets:
                                        if form.is_valid():
                                            form_instance = form.save(commit = False)
                                            form_instance.opening_purchase_voucher_godown_item = shade_form_instance
                            
                                            form_instance.save()
                                    
                messages.success(request,'Item updated successfully')
                return redirect('item-list')
            
        except ProtectedError as e:
            
            messages.error(request,f"Cannot delete item_color_shade due to protected foreign keys: {e}")
            logger.error(f"Cannot delete item_color_shade due to protected foreign keys: {e}")
            print(f"Cannot delete item_color_shade due to protected foreign keys: {e}")

        except Exception as e:
             logger.error(f'An exception occured in item edit - {e}')
             return render(request,'product/item_create_update.html',{'gsts':gsts,
                                                                 'fab_grp':fab_grp,
                                                                 'unit_name':unit_name,
                                                                 'colors':colors,
                                                                 'title':title,
                                                                 'packaging_material_all':packaging_material_all,
                                                                 'fab_finishes':fab_finishes,
                                                                 'form':form,
                                                                 'formset': formset})
        
    return render(request,'product/item_create_update.html',{'gsts':gsts,
                                                                 'fab_grp':fab_grp,
                                                                 'unit_name':unit_name,
                                                                 'colors':colors,
                                                                 'title':title,
                                                                 'packaging_material_all':packaging_material_all,
                                                                 'fab_finishes':fab_finishes,
                                                                 'form':form,
                                                                 'formset': formset})




@login_required(login_url='login')
def openingquantityformsetpopup(request,parent_row_id=None,primary_key=None):
    
    godowns =  Godown_raw_material.objects.all()

    formset = None
    shade_instance = None
    try:
        
        if parent_row_id is not None and primary_key is not None:
            shade_instance = get_object_or_404(item_color_shade,pk=primary_key)
            formset = OpeningShadeFormSetupdate(request.POST or None, instance = shade_instance, prefix = "opening_shade_godown_quantity_set")
            
        
        elif primary_key is None and parent_row_id is not None:

            decoded_data = False
            encoded_data = request.GET.get('data')
            
            
            if encoded_data:
                decoded_data = json.loads(urllib.parse.unquote(encoded_data))
                new_row_data = decoded_data.get('newData', {})
                initial_data_backend = []
                
                
                for key, value in new_row_data.items():
                    initial_data_backend.append({
                            "opening_godown_id": int(value['godownData']),
                            "opening_quantity": float(value['qtyData']),
                            "opening_rate": float(value['rateData'])})

                
                
                total_forms = len(initial_data_backend)
                opening_shade_godown_quantitycreateformset = modelformset_factory(opening_shade_godown_quantity, fields = ['opening_godown_id','opening_quantity','opening_rate'], extra=total_forms, can_delete=True)   
                formset = opening_shade_godown_quantitycreateformset(queryset=opening_shade_godown_quantity.objects.none(),initial=initial_data_backend, prefix = "opening_shade_godown_quantity_set")
                
            else:
                
                opening_shade_godown_quantitycreateformset = modelformset_factory(opening_shade_godown_quantity, fields = ['opening_rate','opening_quantity','opening_godown_id'], extra=1, can_delete=True)            
                formset = opening_shade_godown_quantitycreateformset(queryset=opening_shade_godown_quantity.objects.none(),prefix = "opening_shade_godown_quantity_set")
        
        if request.method == 'POST':
            
            
            if primary_key is not None:

                for form in formset.deleted_forms: 
                    if form.instance.pk:
                        try:
                            form.instance.delete()
                        except Exception as e:

                            return JsonResponse({"error": str(e)}, status=400)

                formset.forms = [form for form in formset if form.has_changed()] 
                if formset.is_valid():
                    for form in formset:
                        if form.is_valid():
                            if not form.cleaned_data.get('DELETE'):
                                try:
                                    old_opening_instance = opening_shade_godown_quantity.objects.get(id= form.instance.id) 
                                    
                                    form_instance = form.save(commit=False)

                                    form_instance.old_opening_godown_id = old_opening_instance.opening_godown_id 
                                    form_instance.old_opening_g_quantity = old_opening_instance.opening_quantity 
                                    form_instance.save()

                                except opening_shade_godown_quantity.DoesNotExist:
                                    form_instance = form.save(commit=False)
                                    
                                    form_instance.old_opening_g_quantity = 0 
                                    form_instance.save()
                                
                                except Exception as e:
                                    
                                    return JsonResponse({"error": str(e)}, status=400)

                    return HttpResponse('<script>window.close();</script>') 
                
                else:
                    logging.error(f'Error in item opening godown formset{formset.error})')
                    return JsonResponse({"errors": formset.errors}, status=400)
                
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)     
    return render(request,'product/opening_godown_qty.html',{'formset':formset,'godowns':godowns ,"parent_row_id":parent_row_id, 'primary_key':primary_key,'shade_instance':shade_instance})




@login_required(login_url='login')
def openingquantityformsetpopupajax(request):
    itemValue_get = request.GET.get('itemValue')
    primary_key_id_get = request.GET.get('primary_key_id')        

    if itemValue_get is not None and primary_key_id_get != '':
        popup_url = reverse('opening-godown-qty-pk', args=[primary_key_id_get,itemValue_get])

    elif itemValue_get is not None:
        popup_url = reverse('opening-godown-qty', args=[itemValue_get])

    else:
        popup_url = None
    
    return JsonResponse({'popup_url':popup_url})




@login_required(login_url='login')
def item_delete(request, pk):
    
    try:
        item_pk = get_object_or_404(Item_Creation,pk = pk)
        item_pk.delete()
        messages.success(request,f'Item {item_pk.item_name} was deleted')
    
    
    except Exception as e:
         messages.error(request, f'EXCEPTION-{e}')
         logger.error( f'EXCEPTION-{e}')
         print(e)
    return redirect('item-list')









@login_required(login_url='login')
def color_create_update(request, pk=None):

    queryset = Color.objects.all()
    color_search = request.GET.get('color_search','')

    if color_search != '':
        queryset = Color.objects.filter(color_name__icontains = color_search)

    if request.path == '/simple_colorcreate_update/':
        template_name = 'product/color_create_update.html'
        title = 'Create Color'

    elif request.path == f'/simple_colorcreate_update/{pk}':
        template_name = 'product/color_create_update.html'
        title = 'Update Color'

    elif request.path == '/colorcreate_update/':
        template_name = "product/create_color_modal.html"
        title = 'Colors'

    elif request.path == '/color_popup/':
        template_name = "product/color_popup.html"
        title = 'Create Colors'
    

    if pk:
        instance = get_object_or_404(Color, pk = pk)
    else:
        instance = None

    form = ColorForm(instance=instance)
    if request.method == 'POST':

        form = ColorForm(request.POST, instance=instance)

        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            
            if request.path == '/simple_colorcreate_update/' or request.path == f'/simple_colorcreate_update/{pk}':
                if instance:
                    messages.success(request, 'Color updated successfully.')
                else:
                    messages.success(request, 'Color created successfully.')
                
                return redirect('simplecolorlist')
            
            elif template_name == "product/color_popup.html":
                
                color_all = Color.objects.all().values('id','color_name')
                messages.success(request, 'Color created successfully.')
                return JsonResponse({'color_all':list(color_all)}) 
        else:
            return render(request, template_name, {'title': title,'form': form,'colors':queryset,'color_search':color_search})
    
    return render(request, template_name , {'title': title, 'form': form, 'colors':queryset,'color_search':color_search})




@login_required(login_url='login')
def color_delete(request, pk):
    try:
        product_color = get_object_or_404(Color,pk=pk)
        product_color.delete()
        messages.success(request,f'Color {product_color.color_name} was deleted')

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {product_color.color_name} because it is referenced by other objects.')
    return redirect('simplecolorlist')










@login_required(login_url='login')
def item_fabric_group_create_update(request, pk = None):
    queryset = Fabric_Group_Model.objects.all()

    fabric_group_search = request.GET.get('fabric_group_search','')

    if fabric_group_search != '':
        queryset = Fabric_Group_Model.objects.filter(fab_grp_name__icontains = fabric_group_search)

    if pk:
        item_fabric_pk =  get_object_or_404(Fabric_Group_Model,pk=pk)
        instance = item_fabric_pk
        title = 'Fabric Group Update'
    else:
        form = ItemFabricGroup()
        instance = None
        title = 'Fabric Group Create'
    

    if request.path == '/itemfabricgroupcreateupdate/':
        template_name = 'product/item_fabric_group_create_update.html'

    elif request.path == '/fabric_popup/':
        template_name = 'product/fabric_popup.html'

    elif request.path == f'/itemfabricgroupcreateupdate/{pk}':
        template_name = 'product/item_fabric_group_create_update.html'

    form = ItemFabricGroup(instance=instance)
    if request.method == 'POST':
        form = ItemFabricGroup(request.POST, instance=instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'Fabric group updated sucessfully.')
            else:
                messages.success(request,'Fabric group created sucessfully.')

            if template_name == 'product/item_fabric_group_create_update.html':
                return redirect('item-fabgroup-create-list')

            elif template_name == 'product/fabric_popup.html':
                Fabric_Group_all = Fabric_Group_Model.objects.all().values('id','fab_grp_name')
                return JsonResponse({'Fabric_Group_all':list(Fabric_Group_all)})

        else:
            
            return render(request,template_name,{'title': title,"fab_group_all":queryset,"fabric_group_search":fabric_group_search,
                                                                                  'form':form})


    return render(request,template_name,{'title': title, "fab_group_all":queryset,"fabric_group_search":fabric_group_search,
                                                                          'form':form})




@login_required(login_url='login')
def item_fabric_group_delete(request,pk):
    try:
        item_fabric_pk = get_object_or_404(Fabric_Group_Model,pk=pk)
        item_fabric_pk.delete()
        messages.success(request,f'Fabric group {item_fabric_pk.fab_grp_name} was deleted')

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {item_fabric_pk.fab_grp_name} because it is referenced by other objects.')
    
    return redirect('item-fabgroup-create-list')








@login_required(login_url='login')
def unit_name_create_update(request,pk=None):
    
    queryset = Unit_Name_Create.objects.all()

    unit_name_search =  request.GET.get('unit_name_search','')

    if unit_name_search != '':
        queryset = Unit_Name_Create.objects.filter(unit_name__icontains = unit_name_search)


    if pk:
        unit_name_pk = get_object_or_404(Unit_Name_Create,pk=pk)
        instance = unit_name_pk
        title = 'Unit Update'
    else:
        instance= None
        title = 'Unit Create'

    form = UnitName(instance = instance)

    if request.path == '/unitnamecreate/':
        template_name = 'product/unit_name_create_update.html'

    elif request.path == '/units_popup/':
        template_name = 'product/units_popup.html'
    
    elif request.path == f'/unitnameupdate/{pk}':
        template_name = 'product/unit_name_create_update.html'

    if request.method == 'POST':
        form = UnitName(request.POST, instance=instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'Unit updated sucessfully.')
            else:
                messages.success(request,'Unit created sucessfully.')

            
            if  template_name == 'product/unit_name_create_update.html':
                return redirect('unit_name-create_list')

            elif template_name == 'product/units_popup.html':
                Unit_Name_all_values = Unit_Name_Create.objects.all().values('id','unit_name')
                return JsonResponse({'Unit_Name_all_values':list(Unit_Name_all_values)})

        else:
            
            return render(request, template_name, {'title': title,'form':form,"unit_name_all":queryset,'unit_name_search':unit_name_search})
        
    else:
        return render(request, template_name, {'title':title,'form':form,"unit_name_all":queryset,'unit_name_search':unit_name_search})




@login_required(login_url='login')
def unit_name_units_ajax(request):
    unit_name_pk = request.GET.get('unit_name_pk')
    if unit_name_pk is not None:
        unit_name_instance = get_object_or_404(Unit_Name_Create,pk=unit_name_pk)
        unit_name_units = unit_name_instance.unit_value
        return JsonResponse({'unit_name_units':unit_name_units})
        



@login_required(login_url='login')
def unit_name_delete(request,pk):
    try:
        unit_name_pk = get_object_or_404(Unit_Name_Create,pk=pk)
        unit_name_pk.delete()
        messages.success(request,f'Unit name {unit_name_pk.unit_name} was deleted.')
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {unit_name_pk.unit_name} because it is referenced by other objects.')
    return redirect('unit_name-create_list')











@login_required(login_url='login')
def account_sub_group_create_update(request, pk=None):

    groups = AccountSubGroup.objects.select_related('acc_grp').all()

    if pk:
        instance = get_object_or_404(AccountSubGroup ,pk=pk)
        title = 'Update'
    else:
        instance = None
        title = 'Create'

    main_grp = AccountGroup.objects.all()
    form = account_sub_grp_form(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Account sub-group created sucessfully')
            return redirect('account_sub_group-create')
        else:
            return render(request,'product/acc_sub_grp_create_update.html', {'main_grp':main_grp,
                                                                             'title':title,
                                                                             'form':form, "groups":groups})

    return render(request,'product/acc_sub_grp_create_update.html', {'main_grp':main_grp, 
                                                                     'title':title,
                                                                     'form':form, "groups":groups})



@login_required(login_url='login')
def account_sub_group_delete(request, pk):  
    try:
        group = get_object_or_404(AccountSubGroup ,pk=pk)
        group.delete()
        messages.success(request,f'Account Sub Group {group.account_sub_group} was deleted')
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {group.account_sub_group} because it is referenced by other objects.')
    return redirect('account_sub_group-create')



@login_required(login_url='login')
def stock_item_create_update(request,pk=None):
    
    if pk:
        instance = get_object_or_404(StockItem ,pk=pk)
        title = 'Stock Item Update'
    else:
        instance = None
        title = 'Stock Item Update'

    if request.user.is_superuser:
        stocks = StockItem.objects.all()
    else:
        stocks = StockItem.objects.filter(company = request.user.company)
   
    accsubgrps = AccountSubGroup.objects.all()

    form = StockItemForm(instance = instance, user = request.user)

    if request.method == 'POST':
        
        form = StockItemForm(request.POST ,instance=instance, user=request.user)
        if form.is_valid():
            try:
                form.save()

                if pk:
                    messages.success(request, 'Stock item updated sucessfully')
                else:
                    messages.success(request, 'Stock item created sucessfully')

                return redirect('stock-item-create')
            except ValidationError as ve:
                messages.error(request, f'{ve}')
            except Exception as e:
                messages.error(request,f'{e}')

        else:
            return render(request,'product/stock_item_create_update.html', {'title':'Stock Item Create',
                                                                            'accsubgrps':accsubgrps,
                                                                            'form':form,'stocks':stocks})
    
    
    return render(request,'product/stock_item_create_update.html', {'title':'Stock Item Create',
                                                                    'accsubgrps':accsubgrps,
                                                                    'form':form,'stocks':stocks})



"""
When to Pass the User:
Role-Based Access: If you want to limit the choices in dropdowns or fields based on the user's role (e.g., superusers can see all companies, while regular users can only see their own), passing the user helps implement this.

Custom Logic: If you have logic in the form's __init__ method that depends on the user (like hiding fields or setting default values), you should pass the user.

Validation: If you need to perform validation based on the user’s company or permissions, having the user available can be useful.

When You Might Not Need to Pass the User:
Simple Forms: If your form doesn’t change based on the user and all users should see the same options, you might not need to pass the user at all.

Static Choices: If the form fields can be populated with static choices or querysets that do not depend on the user's identity, you can simply define those in the form without needing to know who the user is.
"""



@login_required(login_url='login')
def stock_item_delete(request, pk):
    try:
        stock = get_object_or_404(StockItem ,pk=pk)
        stock.delete()
        messages.success(request,f'Stock Item {stock.stock_item_name} was deleted')
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {stock.stock_item_name} because it is referenced by other objects.')        
    return redirect('stock-item-create')



@login_required(login_url='login')
@transaction.atomic
def ledgercreate(request):

    if request.path == '/ledgerpopupcreate/':
        template_name = 'accounts/ledger_create_popup.html'
    else:
        template_name = 'accounts/ledger_create_update.html'

    under_groups = AccountSubGroup.objects.all()
    ledgerTypes_query = ledgerTypes.objects.all()
    form = LedgerForm()
    if request.method == 'POST':
        form = LedgerForm(request.POST)
        
        if form.is_valid():
            ledger_instance = form.save(commit = False) 
            form.save()
            open_bal_value = form.cleaned_data['opening_balance']
            debit_credit_value = form.cleaned_data['Debit_Credit']

            if debit_credit_value == 'Debit':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger' ,debit= open_bal_value)

            elif debit_credit_value == 'Credit':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger',credit= open_bal_value)

            elif debit_credit_value == 'N/A':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger',credit= 0, debit= 0)

            messages.success(request,'Ledger Created')
            
            if request.path == '/ledgerpopupcreate/':
                ledger_labour = Ledger.objects.filter(under_group__account_sub_group='Sundry Creditors').values('id','name')
                              
                return JsonResponse({'ledger_labour':list(ledger_labour)})
            
            else:
                return redirect('ledger-list')
        else:
            return render(request,template_name,{'form':form,'under_groups':under_groups,'title':'ledger Create','ledgerTypes_query':ledgerTypes_query})
    
    return render(request, template_name ,{'form':form,'under_groups':under_groups,'title':'ledger Create','ledgerTypes_query':ledgerTypes_query})
    



@login_required(login_url='login')
@transaction.atomic
def ledgerupdate(request,pk):
    under_groups = AccountSubGroup.objects.all()
    
    Ledger_pk = get_object_or_404(Ledger,pk = pk)
    ledgers = Ledger_pk.transaction_entry.all() 

    Opening_ledger = ledgers.filter(voucher_type ='Ledger').first() 
    form = LedgerForm(instance = Ledger_pk)
    opening_balance = 0

    if form.instance.Debit_Credit == 'Debit':            
        opening_bal = Opening_ledger.debit               
        opening_balance = opening_balance + opening_bal  

    elif form.instance.Debit_Credit == 'Credit':         
        opening_bal = Opening_ledger.credit              
        opening_balance = opening_balance + opening_bal  

    elif form.instance.Debit_Credit == 'N/A': 
        opening_balance = opening_balance 

    else:
        messages.error(request,' Error with Credit Debit ')
    
    if request.method == 'POST':
        
        form = LedgerForm(request.POST, instance = Ledger_pk)
        name_for_message = request.POST['name']
        if form.is_valid():
            form.save()

            if request.POST['Debit_Credit'] == 'Debit':
                Opening_ledger.debit = request.POST['opening_balance']
                Opening_ledger.credit = 0
                Opening_ledger.save()

            if request.POST['Debit_Credit'] == 'Credit':
                Opening_ledger.credit = request.POST['opening_balance']
                Opening_ledger.debit = 0
                Opening_ledger.save()

            if request.POST['Debit_Credit'] == 'N/A':
                Opening_ledger.credit = 0
                Opening_ledger.debit = 0
                Opening_ledger.save()
            
            messages.success(request, f'Ledger of {name_for_message} Updated')
            return redirect('ledger-list')
        else:
            
            return render(request,'accounts/ledger_create_update.html',{'form':form,'under_groups':under_groups,'title':'ledger Update', 'open_bal':opening_balance})
    
    return render(request,'accounts/ledger_create_update.html',{'form':form,'under_groups':under_groups,'title':'ledger Update', 'open_bal':opening_balance})




@login_required(login_url='login')
def ledgerlist(request):
    ledgers = Ledger.objects.select_related('under_group').all()
    return render(request, 'accounts/ledger_list.html', {'ledgers':ledgers})




@login_required(login_url='login')
def ledgerdelete(request, pk):
    try:
        Ledger_pk = get_object_or_404(Ledger ,pk=pk)
        Ledger_pk.delete()
        messages.success(request ,f'ledger of {Ledger_pk.name} was deleted')
    except Exception as e:
        messages.error(request ,f'Cannot delete {Ledger_pk.name} because it is referenced by other objects.')
    return redirect('ledger-list')




@login_required(login_url='login')
def ledgerTypes_create_update(request,pk=None):

    ledger_types = ledgerTypes.objects.all()

    if request.path == '/ledgertypecreatepopup/':
        template_name = 'accounts/ledgertypecreatepopup.html'

    else:
        template_name = 'accounts/ledgerTypescreateupdate.html'

    if pk is not None:
        type_instance = ledgerTypes.objects.get(pk=pk)
    else:
        type_instance = None

    form = ledger_types_form(request.POST or None, instance = type_instance)

    if request.method == 'POST':
        if form.is_valid():
            form.save()

            if request.path == '/ledgertypecreatepopup/':
                ledger_types = ledgerTypes.objects.all().values('id','type_name')
                return JsonResponse({'ledger_type':list(ledger_types)}, status = 200)
            else:  
                return redirect('ledger-Types-create')
            
    return render(request,template_name,{'form':form,'ledger_types':ledger_types})




@login_required(login_url='login')
def ledgerTypes_delete(request,pk):
    type_instance = get_object_or_404(ledgerTypes,pk=pk)
    if type_instance:
        type_instance.delete()
        return redirect('ledger-Types-create')









@login_required(login_url='login')
def godowncreate(request):
    if request.method == 'POST':
       
        godown_name =  request.POST['godown_name']
        godown_type = request.POST['Godown_types']

        if godown_type == 'Raw Material':
            try:
                godown_raw = Godown_raw_material(godown_name_raw=godown_name) 
                
                
                
                

                godown_raw.save()  
                messages.success(request,'Raw material godown created.')

                if 'save' in request.POST:
                    return redirect('godown-list')
                elif 'save_and_add_another' in request.POST:
                    return redirect('godown-create')
                
            except ValidationError as ve:
                messages.error(request,f"{ve}")
                return redirect('godown-list')

            except Exception as e:
                messages.error(request,f"{e}")
                return redirect('godown-list')
        
        elif godown_type == 'Finished Goods':
            try:
                godown_finished = Godown_finished_goods(godown_name_finished=godown_name) 
                

                
                
                

                godown_finished.save() 
                messages.success(request,'Finished goods godown created.')

                if 'save' in request.POST:
                    return redirect('godown-list')
                elif 'save_and_add_another' in request.POST:
                    return redirect('godown-create')
            
            except ValidationError as ve:
                messages.error(request,f"{ve}")
                return redirect('godown-list')

            except Exception as e:
                messages.error(request,f"{e}")
                return redirect('godown-list')
            
        else:
            messages.error(request,'Error Selecting Godown.')
            return redirect('godown-list')
            
    return render(request,'misc/godown_create.html')



    
@login_required(login_url='login')
def godownupdate(request,str,pk):
    if str == 'finished':
        godown_type = 'Finished Goods'
        finished_godown_pk = get_object_or_404(Godown_finished_goods, pk=pk)
        instance_data = finished_godown_pk.godown_name_finished
        if request.method == 'POST':
            godown_name =  request.POST['godown_name']
            finished_godown_pk.godown_name_finished = godown_name
            
            finished_godown_pk.save()
            messages.success(request,'Finished goods godown updated.')
            return redirect('godown-list')
        
    elif str == 'raw':
        godown_type = 'Raw Material'
        raw_godown_pk = get_object_or_404(Godown_raw_material , pk=pk)
        instance_data = raw_godown_pk.godown_name_raw
        if request.method == 'POST':
            godown_name =  request.POST['godown_name']
            raw_godown_pk.godown_name_raw = godown_name
            
            raw_godown_pk.save()
            messages.success(request,'Raw material godown updated.')
            return redirect('godown-list')
    else:
        messages.error(request,'error in godownupdate str variable')
    
    context = {
        'instance_data': instance_data,
        'godown_type': godown_type
    }
    
    return render(request,'misc/godown_update.html', context)




@login_required(login_url='login')
def godownlist(request):

    godowns_raw = Godown_raw_material.objects.all()
    godowns_finished = Godown_finished_goods.objects.all()
    
    
    
        
    
    
    

    return render(request,'misc/godown_list.html',{'godowns_raw':godowns_raw, 
                                                   'godowns_finished':godowns_finished})




@login_required(login_url='login')
def godowndelete(request,str,pk):
    if str == 'finished':
        try:
            finished_godown_pk = get_object_or_404(Godown_finished_goods, pk=pk)
            finished_godown_pk.delete()
            messages.success(request,f'Finished Goods Godown {finished_godown_pk.godown_name_finished} was deleted')
            
        except Exception as e:
            messages.error(request,f'Cannot delete {finished_godown_pk.godown_name_finished} because it is referenced by other objects. ')
            

    elif str == 'raw':
        try:
            raw_godown_pk = get_object_or_404(Godown_raw_material, pk=pk)
            raw_godown_pk.delete()
            messages.success(request,f'Raw Material Godown - {raw_godown_pk.godown_name_raw} was deleted')
            
        except Exception as e:
            messages.error(request,f'Cannot delete {raw_godown_pk.godown_name_raw} because it is referenced by other objects. ')
            
    else:
        messages.error(request, f'Error Deleting Godowns')
    return redirect('godown-list')
    
    







@login_required(login_url='login')
def stockTrasferRaw(request, pk=None):
    godowns = Godown_raw_material.objects.all()

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        
        try:
            
            selected_source_godown_id = request.GET.get('selected_godown_id')
            
            
            destination_godowns_queryset = Godown_raw_material.objects.exclude(pk = selected_source_godown_id)
            
            destination_godowns = {}

            for records in destination_godowns_queryset:
                destination_godowns[records.id] = records.godown_name_raw
            
            selected_source_godown_items = item_godown_quantity_through_table.objects.filter(godown_name=selected_source_godown_id)

            
            items_in_godown = {}
            for items in selected_source_godown_items:
                item = items.Item_shade_name
                item_name =  item.items.item_name
                item_id = item.items.id
                items_in_godown[item_id] = item_name



            
            
            item_name_value = request.GET.get('item_value')
            
            
            item_shades_of_selected_item = item_color_shade.objects.filter(items=item_name_value)

            
            item_shades = {}

            
            items_shade_quantity_in_godown = {}

            
            for x in item_shades_of_selected_item:
                
                shades_of_item_in_selected_godown = item_godown_quantity_through_table.objects.filter(godown_name = selected_source_godown_id, Item_shade_name = x.id)
        
                
                
                for x in shades_of_item_in_selected_godown:

                    shade_name = x.Item_shade_name.item_shade_name
                    shade_id = x.Item_shade_name.id
                    item_shades[shade_id] = shade_name

                    
                    item_id = x.Item_shade_name.id
                    items_shade_quantity_in_godown[item_id] = x.quantity


            
            item_color = None
            item_per = None


            if item_name_value is not None:
                item_name_value = int(item_name_value)

                
                items =  get_object_or_404(Item_Creation ,id = item_name_value)
            
                item_color = items.Item_Color.color_name
                item_per = items.unit_name_item.unit_name


            
            shade_quantity = 0
            selected_shade = request.GET.get('selected_shade_id')
            
            if selected_shade is not None and selected_source_godown_id is not None:
                selected_shade = int(selected_shade)
                selected_source_godown_id = int(selected_source_godown_id)

                quantity_get = item_godown_quantity_through_table.objects.filter(Item_shade_name = selected_shade, godown_name = selected_source_godown_id).first()
                shade_quantity  = quantity_get.quantity
            
            
            return JsonResponse({'items_in_godown': items_in_godown, 'item_shades':item_shades,
                                    'item_color':item_color,'item_per':item_per,'items_shade_quantity_in_godown':items_shade_quantity_in_godown,
                                    'shade_quantity':shade_quantity,'destination_godowns':destination_godowns})
        

        except Exception as e:
            messages.error(request, f'An Error occoured {e}')
            logger.error(f'An Error occoured in stock transfer raw{e}')


    if pk:
        raw_transfer_instance = get_object_or_404(RawStockTransferMaster,voucher_no=pk)
        formset  = raw_material_stock_trasfer_items_formset(request.POST or None, instance = raw_transfer_instance)
        source_godown_items = item_godown_quantity_through_table.objects.filter(godown_name = raw_transfer_instance.source_godown.id)

    else:
        source_godown_items = None
        raw_transfer_instance = None
        formset  = raw_material_stock_trasfer_items_formset(request.POST or None,instance = raw_transfer_instance)


    masterstockform = raw_material_stock_trasfer_master_form(request.POST or None, instance = raw_transfer_instance)


    if request.method == 'POST':
        
        formset.forms = [form for form in formset if form.has_changed()]


        if masterstockform.is_valid() and formset.is_valid():
            masterforminstance = masterstockform.save(commit=False)
            masterforminstance.save()
            
            
            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()

            for form in formset:
                if form.is_valid():
                    if not form.cleaned_data.get('DELETE'):  
                        transfer_instance = form.save(commit=False)
                        transfer_instance.master_instance = masterforminstance 
                        transfer_instance.save()

            return redirect('stock-transfer-raw-list')


    context = {'masterstockform':masterstockform,'formset':formset,'godowns':godowns,'source_godown_items':source_godown_items}

    return render(request,'misc/stock_transfer_raw.html', context=context)



@login_required(login_url='login')
def stockTrasferRawList(request):

    stocktrasferall = RawStockTransferMaster.objects.all()

    return render(request,'misc/stock_transfer_raw_list.html',{'stocktrasferall':stocktrasferall})



@login_required(login_url='login')
def stockTrasferRawDelete(request,pk):
    stocktrasferinstance =  get_object_or_404(RawStockTransferMaster,pk = pk)
    stocktrasferinstance.delete()
    return redirect('stock-transfer-raw-list')








@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def purchasevouchercreateupdate(request, pk = None):
    
    item_name_searched = Item_Creation.objects.all()
    if request.META.get('HTTP_X_REQUESTED_WITH') != 'XMLHttpRequest':

        
        if pk:
            purchase_invoice_instance = get_object_or_404(item_purchase_voucher_master,pk=pk)
            item_formsets_change = purchase_voucher_items_formset_update(request.POST or None, instance=purchase_invoice_instance)
        
        else:
            purchase_invoice_instance = None
            item_formsets_change = purchase_voucher_items_formset(request.POST or None, instance=purchase_invoice_instance)

        items_formset = item_formsets_change
        Purchase_gst = gst.objects.all()

        for forms in items_formset.forms:
            godown_items_formset = purchase_voucher_items_godown_formset()

        raw_material_godowns = Godown_raw_material.objects.all()

        master_form  = item_purchase_voucher_master_form(instance=purchase_invoice_instance)
        
        account_sub_grp = AccountSubGroup.objects.filter(account_sub_group__icontains = 'Sundry Creditors').first()
        
        if account_sub_grp is not None:
            
            party_names = Ledger.objects.filter(under_group=account_sub_grp.id)
        else:
            party_names = ''
    
    try:
        party_gst_no = ''
        selected_party_name = request.GET.get('selected_party_name')
        if selected_party_name is not None:
            ledger_instance = Ledger.objects.filter(id = selected_party_name).first()
            party_gst_no = party_gst_no + ledger_instance.Gst_no

        item_value = request.GET.get('item_value')
        
        item_color_out = ''
        item_per_out = ''
        item_gst_out = 0
        
    
        if item_value is not None: 
            item_value = int(item_value)
            item = Item_Creation.objects.get(id = item_value)

            item_color = item.Item_Color.color_name
            item_color_out =  item_color_out + item_color

            item_per = item.unit_name_item.unit_name
            item_per_out = item_per_out + item_per

            item_gst = item.Item_Creation_GST.gst_percentage
            item_gst_out = item_gst_out + item_gst

        
        
        item_shades = item_color_shade.objects.filter(items = item_value)

        item_shades_dict = {}
        item_shades_total_quantity_dict = {}
        
        for shade in item_shades:
            item_shades_dict[shade.id] = shade.item_shade_name
            
            godown_shade_quantity = 0
            shade_godowns =  item_godown_quantity_through_table.objects.filter(Item_shade_name = shade)
            for godown in shade_godowns:
                godown_shade_quantity = godown_shade_quantity + godown.quantity
            item_shades_total_quantity_dict[shade.id] = godown_shade_quantity
        
        auto_popup_flag = False

        shade_count = len(item_shades_total_quantity_dict)

        if shade_count == 1:
            auto_popup_flag = True


    except Exception as e:
        print(f'exception occoured {e}')
    
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
             return JsonResponse({'item_color': item_color_out , 'item_shade': item_shades_dict,'auto_popup_flag':auto_popup_flag,
                                  "item_per":item_per_out, 'item_shades_total_quantity_dict':item_shades_total_quantity_dict,
                                  'item_gst_out':item_gst_out,'party_gst_no':party_gst_no})

    if request.method == 'POST':
        
        try:
            with transaction.atomic(): 
                
                master_form = item_purchase_voucher_master_form(request.POST,instance=purchase_invoice_instance)

                
                items_formset = item_formsets_change

                
                godown_items_formset = purchase_voucher_items_godown_formset(request.POST, prefix='shade_godown_items_set')

                
                items_formset.forms = [form for form in items_formset.forms] 

                if master_form.is_valid() and items_formset.is_valid():

                    
                    master_instance = master_form.save()

                    
                    
                    
                    for form in items_formset.deleted_forms:
                        
                        if form.instance.pk:
                            
                            form.instance.deleted_directly = True
                            form.instance.delete()

                    all_purchase_temp_data = []
                    
                    for form in items_formset:
                        if form.is_valid():

                            
                            if not form.cleaned_data.get('DELETE'):
                                items_instance = form.save(commit=False)
                                items_instance.item_purchase_master = master_instance
                                items_instance.save()
                                
                                form_prefix_number = form.prefix[-1] 
                                
                                
                                unique_id_no = request.POST.get(f'item_unique_id_{form_prefix_number}')
                                primary_key = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-id')
                                
                                
                                if primary_key == '' or primary_key == None: 
                                    purchase_voucher_temp_data = shade_godown_items_temporary_table.objects.filter(unique_id=unique_id_no)
                                    for data in purchase_voucher_temp_data:
                                        all_purchase_temp_data.append(data)
                                
                                    godown_temp_data = {}
                                    form_set_id = 0
                                    for temp_godown_row_data in purchase_voucher_temp_data:
                                        godown_temp_data[f'shade_godown_items_set-TOTAL_FORMS'] = str(len(purchase_voucher_temp_data))
                                        godown_temp_data[f'shade_godown_items_set-INITIAL_FORMS'] =  str(0)
                                        godown_temp_data[f'shade_godown_items_set-MIN_NUM_FORMS'] =  str(0)
                                        godown_temp_data[f'shade_godown_items_set-MAX_NUM_FORMS'] =  str(1000)
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-godown_id'] = temp_godown_row_data.godown_id
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-quantity'] = temp_godown_row_data.quantity
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-rate'] = temp_godown_row_data.rate
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-amount'] = temp_godown_row_data.amount
                                        form_set_id =  form_set_id + 1
                                    
                                    godown_items_formset = purchase_voucher_items_godown_formset(godown_temp_data, prefix='shade_godown_items_set')
                                    saved_data_to_delete = 0
                                    for godown_form in godown_items_formset:
                                        if godown_form.is_valid():
                                            godown_instance = godown_form.save(commit = False)
                                            godown_instance.purchase_voucher_godown_item = items_instance 
                                            godown_instance.save()
                                            saved_data_to_delete = saved_data_to_delete + 1
                                            
                                        else:
                                           
                                            purchase_voucher_temp_data.delete()

                                    if saved_data_to_delete == form_set_id:
                                        purchase_voucher_temp_data.delete()


                                
                                godown_item_quantity = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-jsonDataInputquantity')
                               
                                if godown_item_quantity != '':
                                    voucher_row_godown_data = json.loads(godown_item_quantity)
                                    
                                    parent_row_prefix_id = voucher_row_godown_data.get('parent_row_prefix_id')

                                    if parent_row_prefix_id == form_prefix_number:

                                        new_row = voucher_row_godown_data.get('newRow')
                                        new_rate = float(voucher_row_godown_data.get('all_Rate'))
                                        row_item = items_instance.item_shade.id
                                        Item_instance =  item_color_shade.objects.get(id = row_item)

                                        for key, value in new_row.items():
                                            godown_id = int(value['gId'])    
                                            updated_quantity = value['jsonQty']   
                                            
                                            
                                            godown_old_id = value.get('popup_old_id', None)   
                                            if godown_old_id == '':
                                                godown_old_id = None

                                            if godown_old_id != '' and godown_old_id is not None:
                                                godown_old_id = int(godown_old_id)   

                                            popup_row_id = value.get('popup_row_id', None)  
                                            if popup_row_id == '':
                                                popup_row_id = None
                                        
                                            
                                            if godown_old_id == None or godown_old_id == godown_id:
                                                
                                                godown_instance = Godown_raw_material.objects.get(id = godown_id)
                                                Item, created = item_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance,Item_shade_name = Item_instance)
                                                
                                                if popup_row_id == None or popup_row_id == '' :
                                                    initial_quantity = 0

                                                else:
                                                    initial_quantity = shade_godown_items.objects.get(pk = popup_row_id)
                                                    initial_quantity = initial_quantity.quantity
                                                
                                                qty_to_update = updated_quantity - initial_quantity

                                                Item.quantity = Item.quantity + qty_to_update
                                                Item.item_rate = new_rate
                                                Item.save()
                                            
                                            
                                            if godown_old_id != None:
                                                
                                                godown_old_id = int(godown_old_id) 
                                                godown_instance_old = Godown_raw_material.objects.get(id = godown_old_id)

                                                godown_new_id = int(godown_id)
                                                godown_instance_new = Godown_raw_material.objects.get(id = godown_new_id)

                                                
                                                if godown_old_id != godown_new_id:
                                                    old_quantity_get = shade_godown_items.objects.get(pk = popup_row_id)
                                                    old_quantity = old_quantity_get.quantity

                                                    old_godown_through_row = item_godown_quantity_through_table.objects.get(godown_name = godown_instance_old,Item_shade_name=Item_instance)

                                                    old_godown_through_row.quantity = old_godown_through_row.quantity - old_quantity
                                                    old_godown_through_row.save()

                                                    new_godown_through_row, created  = item_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance_new,Item_shade_name=Item_instance)
                                                    
                                                    
                                                    if new_godown_through_row:
                                                        new_quantity_c = new_godown_through_row.quantity
                                                    else:
                                                        
                                                        new_quantity_c = 0

                                                    new_godown_through_row.quantity = new_quantity_c + updated_quantity
                                                    new_godown_through_row.save()


                                
                                popup_godowns_exists = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-popupData')
                                old_item_shade = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-old_item_shade')
                                
                                if popup_godowns_exists != '':
                                    popup_godown_data = json.loads(popup_godowns_exists)
                                
                                    row_prefix_id = popup_godown_data.get('prefix_id')

                                    if row_prefix_id == form_prefix_number:
                                        
                                        shade_id = int(popup_godown_data.get('shade_id'))
                                        prefix_id =  int(popup_godown_data.get('prefix_id'))
                                        primarykey = int(popup_godown_data.get('primary_id'))
                                        old_item_shade = int(old_item_shade)
                                        
                                        
                                        purchasevoucherpopupupdate(popup_godown_data,shade_id,prefix_id,primarykey,old_item_shade)
                                        
                    return redirect('purchase-voucher-list')
                else:
                    
                    if 'temp_data_exists' in request.session and 'temp_uuid' in request.session: 
                        temp_data_exists_bool = request.session['temp_data_exists']
                        temp_uuids = request.session['temp_uuid']
                        del request.session['temp_data_exists']
                        del request.session['temp_uuid']
                        for data in temp_uuids:
                            temp_uuids_data =  shade_godown_items_temporary_table.objects.filter(unique_id=data)
                            temp_uuids_data.delete()

                    return redirect('purchase-voucher-list')
            
        except Exception as e:
            print('an error occoured-test',e)
            messages.error(request,f'An error occoured{e} godown temporary data deleted')
            
        finally:
                
                
                
                if 'temp_data_exists' in request.session and 'temp_uuid' in request.session: 
                    temp_data_exists_bool = request.session['temp_data_exists']
                    temp_uuids = request.session['temp_uuid']
                    del request.session['temp_data_exists']
                    del request.session['temp_uuid']
                    for data in temp_uuids:
                        temp_uuids_data = shade_godown_items_temporary_table.objects.filter(unique_id=data)
                        temp_uuids_data.delete()

    context = {'master_form':master_form,
               'party_names':party_names,
               'items_formset':items_formset,
               'Purchase_gst':Purchase_gst,
               'godown_formsets':godown_items_formset,
               'item_godowns_raw':raw_material_godowns,
               'items':item_name_searched
               }

    return render(request,'accounts/purchase_invoice.html',context=context)




@login_required(login_url='login')
def purchasevoucherpopupupdate(popup_godown_data,shade_id,prefix_id,primarykey,old_item_shade):
        
        if primarykey is not None:
            voucher_item_instance = purchase_voucher_items.objects.get(id=primarykey)

            if old_item_shade != shade_id:
                all_godown_old_instances = shade_godown_items.objects.filter(purchase_voucher_godown_item = primarykey)
                if all_godown_old_instances:
                    for items in all_godown_old_instances:
                        items.deleted_directly = True

                        
                        items.extra_data_old_shade = old_item_shade
                        items.delete()

            formset = purchase_voucher_items_godown_formset(popup_godown_data, instance = voucher_item_instance ,prefix='shade_godown_items_set')
            
            if formset.is_valid():
                for form in formset.deleted_forms:
                    if form.instance.pk:
                        
                        form.instance.deleted_directly = True
                        form.instance.delete()
                formset.save()
            else:
                print('godown_errors',formset.errors)

                


@login_required(login_url='login')
def purchasevoucherpopup(request,shade_id,prefix_id,unique_id=None,primarykey=None,item_rate=None):
    
    if item_rate != None:
        item_rate_value = decimal.Decimal(item_rate)
    else:
        item_rate_value = None
    
   
    
    
    
    
    if unique_id is not None:
        
        temp_instances = shade_godown_items_temporary_table.objects.filter(unique_id=unique_id)
        
        if temp_instances:
            formsets = shade_godown_items_temporary_table_formset_update(request.POST or None, queryset = temp_instances,prefix='shade_godown_items_set')

        else:
            formsets = shade_godown_items_temporary_table_formset(request.POST or None, queryset = temp_instances,prefix='shade_godown_items_set')
    
    
    elif primarykey is not None:

        godowns_for_selected_shade = shade_godown_items.objects.filter(purchase_voucher_godown_item__item_shade = shade_id,purchase_voucher_godown_item = primarykey)
        
        voucher_item_instance = purchase_voucher_items.objects.get(id=primarykey)
        
        if godowns_for_selected_shade:
            formsets = purchase_voucher_items_godown_formset(instance = voucher_item_instance,prefix='shade_godown_items_set')
        else:
            formsets = purchase_voucher_items_godown_formset_shade_change(
            instance=voucher_item_instance,
            prefix='shade_godown_items_set',
            queryset=godowns_for_selected_shade)

    
    formset = formsets

    try:
        godowns = Godown_raw_material.objects.all()
        item = Item_Creation.objects.get(shades__id = shade_id) 
        item_shade = item_color_shade.objects.get(id = shade_id)

    except Exception as e:
        messages.error(request,'Error with Shades')

    
    if request.method == 'POST':
        formset = formsets

        if formset.is_valid():

            for form in formset.deleted_forms:
                if form.instance.id:
                    form.instance.delete()

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    if form.is_valid():
                        form.save()
                    
            
            
            request.session['temp_data_exists'] = True
            temp_uuid = request.session.get('temp_uuid', [])
            temp_uuid.append(unique_id)
            request.session['temp_uuid'] = temp_uuid 

            return HttpResponse('<script>window.close();</script>') 

        else:
            print(formset.errors)
            print(formset.non_form_errors())
            context = {
                'godowns': godowns, 'item': item, 'item_shade': item_shade, 'formset': formset, 'item_rate_value':item_rate_value,
                'unique_id': unique_id, 'shade_id': shade_id, 'errors': formset.errors,'prefix_id':prefix_id, 'primary_key':primarykey
            }
            return render(request, 'accounts/purchase_popup.html', context)
        
    return render(request, 'accounts/purchase_popup.html' ,{'godowns':godowns,'item':item,'shade_id': shade_id,
                                                            'item_shade':item_shade,'formset':formset,'item_rate_value':item_rate_value,
                                                            'unique_id':unique_id,'prefix_id':prefix_id, 'primary_key':primarykey})




@login_required(login_url='login')
def purchasevouchercreategodownpopupurl(request):
    shade_id = request.GET.get('selected_shade')
    unique_id = request.GET.get('unique_invoice_row_id')
    primary_key = request.GET.get('purchase_id')
    prefix_id  = request.GET.get('prefix_id')
    item_instance = item_color_shade.objects.get(id=shade_id)
    fab_grp_instance = Fabric_Group_Model.objects.get(items__shades = shade_id)

    
    query_set_order = item_color_shade.objects.filter(items__Fabric_Group=fab_grp_instance).order_by('-modified_date_time').first()

    item_rate = query_set_order.rate



    
    if primary_key is not None:
        popup_url = reverse('purchase-voucher-popup-update', args=[shade_id,prefix_id,primary_key])
        
    elif unique_id is not None:
        popup_url = reverse('purchase-voucher-popup-create', args=[shade_id,prefix_id,item_rate,unique_id])
        
    else:
        popup_url = None
    

    return JsonResponse({'popup_url':popup_url})




@login_required(login_url='login')
def purchasevoucherlist(request):
    purchase_invoice_list = item_purchase_voucher_master.objects.all()
    return render(request,'accounts/purchase_invoice_list.html',{'purchase_invoice_list':purchase_invoice_list})




@login_required(login_url='login')
def purchasevoucherdelete(request,pk):
    purchase_invoice_pk = get_object_or_404(item_purchase_voucher_master,pk=pk)
    purchase_invoice_pk.delete()
    return redirect('purchase-voucher-list')
                    









@login_required(login_url='login')
def salesvouchercreate(request):
    return render(request,'.html')




@login_required(login_url='login')
def salesvoucherupdate(request,pk):
    return render(request,'.html')




@login_required(login_url='login')
def salesvoucherlist(request):
    return render(request,'.html')



@login_required(login_url='login')
def salesvoucherdelete(request,pk):
    pass










@login_required(login_url='login')
def gst_create_update(request, pk = None):

    queryset =  gst.objects.all()

    gst_search = request.GET.get('gst_search','')

    if gst_search != "":
        queryset =  gst.objects.filter(gst_percentage__contains = gst_search)


    if pk:
        instance = gst.objects.get(pk=pk)
        title = 'Update'

    else:
        instance = None
        title = 'Create'


    if request.path == '/gstpopup/':
        template_name = 'accounts/gst_popup.html'
    
    elif request.path == '/gstcreate/':
        template_name = 'accounts/gst_create_update.html'


    elif request.path == f'/gstupdate/{pk}':
        template_name = 'accounts/gst_create_update.html'
 

    form = gst_form(instance = instance)
    if request.method == 'POST':
        form = gst_form(request.POST, instance = instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()
            
            if pk:
                messages.success(request,'GST updated successfully.')
            else:
                messages.success(request,'GST created successfully.')

            if template_name == 'accounts/gst_create_update.html':
                
                return redirect('gst-create-list')

            elif template_name == 'accounts/gst_popup.html':
                
                gst_updated = gst.objects.all().values('id', 'gst_percentage')
                
                return JsonResponse({"gst_updated": list(gst_updated)})
        else:
            return render(request,template_name,{'form':form, 'title':title, 'gsts':queryset})

    return render(request,template_name,{'form':form, 'title':title, 'gsts':queryset})




@login_required(login_url='login')
def gst_delete(request,pk):
    gst_pk = gst.objects.get(pk=pk)
    gst_pk.delete()
    messages.success(request,'GST deleted')
    return redirect('gst-create-list')




@login_required(login_url='login')
def fabric_finishes_create_update(request, pk = None):
    queryset =  FabricFinishes.objects.all()

    fabric_finishes_search = request.GET.get('fabric_finishes_search','')

    if fabric_finishes_search != '':
        queryset =  FabricFinishes.objects.filter(fabric_finish__icontains = fabric_finishes_search)

    if pk:
        fabric_finishes_instance = FabricFinishes.objects.get(pk=pk)
        title = 'Update'
    else:
        fabric_finishes_instance = None
        title = 'Create'

    if request.path == '/fabricfinishespopup/':
        template_name = 'misc/fabric_finishes_popup.html'

    elif request.path == '/fabricfinishesscreate/' or f'/fabricfinishesupdate/{pk}':
        template_name = 'misc/fabric_finishes_create_update.html'

    form = FabricFinishes_form(instance = fabric_finishes_instance)

    if request.method == 'POST':
        form = FabricFinishes_form(request.POST,instance = fabric_finishes_instance)

        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'fabric finish updated sucessfully')
            else:
                messages.success(request,'fabric finish created sucessfully')

            if template_name == 'misc/fabric_finishes_create_update.html':
                return redirect('fabric-finishes-create-list')
            
            elif template_name == 'misc/fabric_finishes_popup.html':
                fabric_finishes_all = FabricFinishes.objects.all().values('id', 'fabric_finish')
                
                return JsonResponse({"fabric_finishes_all": list(fabric_finishes_all)})
        else:
            
            return render(request,template_name,{'form':form,'title':title,'fabricfinishes':queryset,'fabric_finishes_search':fabric_finishes_search})

    return render(request,template_name,{'form':form,'title':title,'fabricfinishes':queryset,'fabric_finishes_search':fabric_finishes_search})




@login_required(login_url='login')
def fabric_finishes_delete(request,pk):
    fabric_finish =  FabricFinishes.objects.get(pk=pk)
    fabric_finish.delete()
    messages.success(request,'fabric finish deleted.')
    return redirect('fabric-finishes-create-list')



@login_required(login_url='login')
def packaging_create_update(request, pk = None):
    
    queryset =  packaging.objects.all()

    packaging_search = request.GET.get('packaging_search','')

    if packaging_search != '':
        queryset =  packaging.objects.filter(packing_material__icontains = packaging_search)

    if pk:
        packaging_instance = packaging.objects.get(pk=pk)
        title = 'Update'
    else:
        packaging_instance = None
        title = 'Create'

    if request.path == '/packagingpop/':
        template_name = 'misc/packaging_popup.html'

    elif request.path == '/packaging_create/' or f'/packagingupdate/{pk}':
        template_name = 'misc/packaging_create_update.html'

    form = packaging_form(instance = packaging_instance)

    if request.method == 'POST':
        form = packaging_form(request.POST ,instance=packaging_instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'packing updated sucessfully.')
            else:
                messages.success(request,'packing created sucessfully.')

            
            if template_name == 'misc/packaging_create_update.html':
                return redirect('packaging-create-list')

            elif template_name == 'misc/packaging_popup.html':

                packaging_all_values = packaging.objects.all().values('id','packing_material')

                return JsonResponse({'packaging_all_values': list(packaging_all_values)})
        else:
            return render(request, template_name ,{'form':form,'title':title,'packaging_all':queryset}) 

    return render(request, template_name ,{'form':form,'title':title,'packaging_all':queryset})




@login_required(login_url='login')
def packaging_delete(request,pk):
    packaging_pk =  packaging.objects.get(pk=pk)
    packaging_pk.delete()
    messages.success(request, 'Packing deleted.')
    return redirect('packaging-create-list')










@login_required(login_url='login')
def product2item(request,product_refrence_id):
    
    try:
        items = Item_Creation.objects.all().order_by('item_name')
        product_refrence_no = product_refrence_id
        Products_all = PProduct_Creation.objects.filter(Product__Product_Refrence_ID=product_refrence_id).select_related('PProduct_color')

        if not Products_all.exists():
                raise ValueError("No products found for the given reference ID.")
        
        
        extraformspecial = True
        for product in Products_all:
            if product.product_2_item_through_table_set.filter(common_unique = False):
                extraformspecial = False

        
        extraformcommon = True
        for product in Products_all:
            if product.product_2_item_through_table_set.filter(common_unique = True):
                extraformcommon = False
        

        
        
        product2item_instances = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_refrence_id,
              common_unique = False).select_related('PProduct_pk','Item_pk','PProduct_pk__PProduct_color').order_by('row_number')
        

        if extraformspecial:
            formset_single = Product2ItemFormsetExtraForm(queryset = product2item_instances, prefix='product2itemuniqueformset')

        else:
            formset_single = Product2ItemFormset(queryset=product2item_instances , prefix = 'product2itemuniqueformset')

        
        
        
        
        
        
        distinct_product2item_commmon_instances = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_refrence_id,common_unique = True).order_by(
                'row_number','id').distinct('row_number').select_related('PProduct_pk','Item_pk')


        if extraformcommon:
            formset_common = Product2CommonItemFormSetExtraForm(queryset=distinct_product2item_commmon_instances,prefix='product2itemcommonformset')

        else:
            formset_common = Product2CommonItemFormSet(queryset=distinct_product2item_commmon_instances,prefix='product2itemcommonformset')

        if request.method == 'POST':
            
            formset_single = Product2ItemFormset(request.POST, queryset=product2item_instances, prefix='product2itemuniqueformset')
            formset_common = Product2CommonItemFormSet(request.POST, queryset=distinct_product2item_commmon_instances, prefix='product2itemcommonformset') 
            
            formset_single_valid = False
            formset_common_valid = False
            
            
            if formset_single.is_valid():
                
                try:
                    
                    for form in formset_single.deleted_forms:
                        if form.instance.pk:  
                            
                            
                            
                            
                            
                            
                            form.instance.delete()
                            
                            

                    
                    for form in formset_single:
                        if not form.cleaned_data.get('DELETE'): 
                            if form.cleaned_data.get('Item_pk'):  
                                
                                if form.instance.pk:  
                                    existing_instance = product_2_item_through_table.objects.get(pk=form.instance.pk)  
                                    initial_rows = existing_instance.no_of_rows 
                                else:
                                    initial_rows = 0

                                p2i_instance = form.save(commit = False)
                                p2i_instance.c_user = request.user
                                p2i_instance.common_unique = False 
                                p2i_instance.save()
                                logger.info(f"Product to item created/updated special - {p2i_instance.id}")

                                no_of_rows_to_create = form.cleaned_data['no_of_rows'] - initial_rows   
                                p2i_instance.row_number = form.cleaned_data['row_number']

                                if no_of_rows_to_create > 0:
                                    for row in range(no_of_rows_to_create):
                                        
                                        logger.info(f" set prod item part name created of p2i instance - {p2i_instance.id}")
                                        set_prod_item_part_name.objects.create(producttoitem = p2i_instance, c_user = request.user)

                                p2i_instance.save()
                                formset_single_valid = True

                            else:
                                raise ValidationError('Please select existing Item Name or select from the dropdown')
                                
                except Exception as e:
                    logger.error(f'Error saving unique records - {e}')
                    messages.error(request, f'Error saving unique records - {e}')  
            
            else:
                logger.error(f'Error saving unique records - {formset_single.errors}')
                messages.error(request, f'Error saving unique records - {formset_single.errors}') 
                            
            
            if formset_common.is_valid():
                try:
                    for form in formset_common.deleted_forms:
                        if form.instance.id: 
                            deleted_item = form.instance.Item_pk  
                            
                            

                            
                            for product in Products_all: 
                                p2i_to_delete = product_2_item_through_table.objects.filter(PProduct_pk=product, Item_pk=deleted_item, common_unique=True)
                                logger.info(f"Deleted product to item instace of {product}, - {deleted_item}")
                                p2i_to_delete.delete()
                            
                            

                            
                    for form in formset_common: 
                        if not form.cleaned_data.get('DELETE'): 

                            if form.cleaned_data.get('Item_pk'):  

                                for product in Products_all:
                                    
                                    
                                    item = form.cleaned_data['Item_pk']
                                    
                                    obj, created = product_2_item_through_table.objects.get_or_create(PProduct_pk=product, Item_pk=item, common_unique=True)
                                    obj.c_user = request.user
                                    
                                    
                                    if created:
                                        initial_rows = 0

                                    if not created:
                                        initial_rows = obj.no_of_rows
                                    
                                    obj.no_of_rows = form.cleaned_data['no_of_rows']
                                    obj.Remark = form.cleaned_data['Remark']
                                    obj.row_number = form.cleaned_data['row_number']
                                    logger.info(f"Product to item created/updated common - {obj.id}")
                                    obj.save()
                                
                                    
                                    rows_to_create = form.cleaned_data['no_of_rows'] - initial_rows
                                    if rows_to_create > 0:
                                            for row in range(rows_to_create):
                                                set_prod_item_part_name.objects.create(producttoitem = obj,c_user = request.user)
                                                logger.info(f" set prod item part name created of - {obj.id}")

                                    formset_common_valid = True

                            else:
                                raise ValidationError('Please select existing Item Name or select from the dropdown')
                                    

                except Exception as e:
                    logger.error(f'Error saving common records - {e}')
                    messages.error(request, f'Error saving common records{e}.') 

            else:
                logger.error(f'Error saving unique records - {formset_common.errors}')
                messages.error(request, f'Error saving unique records - {formset_common.errors}')
        

            if formset_common_valid and formset_single_valid:

                messages.success(request,'Items to Product sucessfully added.')
                close_window_script = """
                                                <script>
                                                window.opener.location.reload(true);  // Reload parent window if needed
                                                window.close();  // Close current window
                                                </script>
                                                """
                return HttpResponse(close_window_script)
            
            else:
                for form_errors in formset_common.errors:
                    if form_errors:
                        logger.error(f'Error with formset_common form - {product_refrence_id } - {form_errors}')
                        messages.error(request, f'{form_errors}')

                for form_errors in formset_single.errors:
                    if form_errors:
                        logger.error(f'Error with formset_common form - {product_refrence_id } - {form_errors}')
                        messages.error(request, f'{form_errors}')


                close_window_script = """
                            <script>
                                                window.opener.location.reload(true);  // Reload parent window if needed
                                                window.close();  // Close current window
                                                </script>
                                                """
                return HttpResponse(close_window_script)


        return render(request,'production/product2itemsetproduction.html', { 'formset_single':formset_single,'formset_common':formset_common,
                                                                'Products_all':Products_all,
                                                                'items':items,'product_refrence_no': product_refrence_no})

    except Exception as e:
        logger.error(f'Error with forms - {product_refrence_id } - {e}')
        messages.error(request, 'An unexpected error occurred. Please try again later.')
        return render(request, 'production/product2itemsetproduction.html', {
            'formset_single': formset_single,
            'formset_common': formset_common,
            'Products_all': Products_all,
            'items': items,
            'product_refrence_no': product_refrence_no
        })



@login_required(login_url='login')
def export_Product2Item_excel(request,product_ref_id):
    
    try:
        
        products_in_i2p_special = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_ref_id, common_unique = False).order_by(
            'row_number')
    
        products_in_i2p_common = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_ref_id, common_unique = True).order_by(
            'row_number', 'id').distinct('row_number')

        
        if not products_in_i2p_special and not products_in_i2p_common:
            logger.error(f"Add Items to the respective Products first - {product_ref_id}")
            raise ObjectDoesNotExist("Add Items to the respective Products first")

        wb = Workbook()

        
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)    

        wb.create_sheet('product_special_configs')
        wb.create_sheet('product_common_configs')

        sheet1 = wb.worksheets[0]
        sheet2 = wb.worksheets[1]

        column_widths = [10, 40, 20, 30, 20, 15, 12 ,12, 12, 12]  

        
        for i, column_width in enumerate(column_widths, start=1):  
            col_letter = get_column_letter(i)
            sheet1.column_dimensions[col_letter].width = column_width


        
        for i, column_width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            sheet2.column_dimensions[col_letter].width = column_width

        
        headers =  ['id','item name', 'product sku','part name', 'part dimention','dimention total','part pieces','body/combi','grand_total', 'combi_total']
        sheet1.append(headers)

        row_count_to_unlock_total = 1
        for product in products_in_i2p_special:
            grand_total_parent = product.grand_total
            grand_total_combi_parent = product.grand_total_combi

            rows_to_insert_s1 = []

            for product_configs in product.product_item_configs.all().order_by('id'):
                rows_to_insert_s1.append([
                product_configs.id,
                product_configs.producttoitem.Item_pk.item_name,
                product_configs.producttoitem.PProduct_pk.PProduct_SKU,
                product_configs.part_name,
                product_configs.part_dimentions,
                product_configs.dimention_total,
                product_configs.part_pieces,
                product_configs.body_combi
            ])

            row_count_to_unlock = 1

            for row in rows_to_insert_s1:
                sheet1.append(row)

            
            
            
            

                row_count_to_unlock = row_count_to_unlock + 1

            row_count_to_unlock_total =  row_count_to_unlock_total + row_count_to_unlock

            
            sheet1.append(['','','','','','','','', grand_total_parent , grand_total_combi_parent])
        
            rows_to_insert_s1.clear()

        
        for row in sheet1.iter_rows(min_row=2, max_row=row_count_to_unlock_total, min_col=4, max_col=8):
            for cell in row:
                cell.protection = Protection(locked = False)

        
        headers =  ['id','item name','part name', 'part dimention', 'dimention total','part pieces','body/combi','grand_total','combi_total']
        sheet2.append(headers)

        row_count_to_unlock_total_common = 1
        for product in products_in_i2p_common:
            grand_total_parent = product.grand_total
            grand_total_combi_parent = product.grand_total_combi

            rows_to_insert_s2 = []
            for product_configs in product.product_item_configs.all().order_by('id'):
                rows_to_insert_s2.append([
                product_configs.id,
                product_configs.producttoitem.Item_pk.item_name,
                product_configs.part_name,
                product_configs.part_dimentions,
                product_configs.dimention_total,
                product_configs.part_pieces,
                'body',
            ])

            row_count_to_unlock = 1

            for row in rows_to_insert_s2:
                sheet2.append(row)
                row_count_to_unlock = row_count_to_unlock + 1
            row_count_to_unlock_total_common = row_count_to_unlock_total_common + row_count_to_unlock

            
            sheet2.append(['','','','','','','',grand_total_parent, grand_total_combi_parent])

            rows_to_insert_s2.clear()

        
        for row in sheet2.iter_rows(min_row=2, max_row=row_count_to_unlock_total_common, min_col=3, max_col=7):
            for cell in row:
                cell.protection = Protection(locked = False)

        
        sheet1.protection.sheet = True
        sheet2.protection.sheet = True

        fileoutput = BytesIO()
        wb.save(fileoutput)
        
        
        response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name_with_pk = f'product_reference_id_{product_ref_id}'
        response['Content-Disposition'] = f'attachment; filename="{file_name_with_pk}.xlsx"'

        return response
    
    except product_2_item_through_table.DoesNotExist:
        messages.error(request, 'No data found for the given product reference ID.')
        logger.error(f"items for product Does not exists for refrence id {product_ref_id}")
        return redirect(reverse('edit_production_product', args=[product_ref_id]))
    
    
    except Exception as e:
        messages.error(request, f'An error occurred while exporting data {e}')
        logger.error(f"An exception occoured while exporting data - {e} for ref id {product_ref_id}")
        return redirect(reverse('edit_production_product', args=[product_ref_id]))



@login_required(login_url='login')
def viewproduct2items_configs(request, product_sku):

    try:
        product2item_instances = product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU=product_sku).order_by('row_number')
        product2item_instances_first = product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU=product_sku).first()


        context = {
            'product2item_instances': product2item_instances,
            'product2item_instances_first': product2item_instances_first
        }

        return render(request, 'production/product2itemsconfigview.html', context)

    except product_2_item_through_table.DoesNotExist:
        return render(request, 'production/product2itemsconfigview.html', {
            'error_message': f'No items found for product SKU: {product_sku}'})

    except DatabaseError as e:
        
        return HttpResponseServerError(f'A database error occurred: {e}')

    except Exception as e:
        
        return HttpResponseServerError(f'An unexpected error occurred: {e}')
    

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseordercreateupdate(request,pk=None):
    
    try:
        ledger_party_names = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Debtors')
        products = Product.objects.all()

        raw_material_godowns = Godown_raw_material.objects.all()

        
        if pk:
            instance = get_object_or_404(purchase_order,pk=pk)
            model_name = instance.product_reference_number.Model_Name
            model_images = instance.product_reference_number.productdetails

        else:
            instance = None
            model_name = None
            model_images = None
        
        formset = purchase_order_product_qty_formset(instance=instance)
        form = purchase_order_form(instance=instance)

    except Exception as e:
        logger.error(f'An Exception Occoured {e}')

    
    if request.method == 'POST':

        
        
        if 'submit-form-1' in request.POST:
            form = purchase_order_form(request.POST, instance=instance)
        
            if form.is_valid():
                try:
                    form_instance = form.save(commit=False)
                    form_instance.balance_number_of_pieces = form.instance.number_of_pieces
                    form_instance.save()
                    logger.info(f'Purchase invoice created-updated-{form.instance.id}')
                    
                    return redirect(reverse('purchase-order-update', args=[form.instance.id]))
                
                except ValidationError as val_err:
                    logger.error(f'Validation error: {val_err} - {form.errors}')

                except DatabaseError as db_err:
                    logger.error(f'Database error during formset save: {db_err}')

                except Exception as e:
                    logger.error(f'Unexpected error during form save: {e}')
            else:
                logger.error(f'Purchase Order Quantities updated error-{form.instance.id} - {form.errors}')
            

        if 'submit-form-2' in request.POST:
            
            try:
                formset = purchase_order_product_qty_formset(request.POST, instance=instance)
                

                
                
                
                

                
                if formset.is_valid():
                    try:
                        formset.save()
                        
                        for form in formset:
                            if not form.cleaned_data.get('DELETE'):
                                p_o_instance = form.instance.purchase_order_id  
                                p_o_instance.purchase_order_to_product_saved = True
                                p_o_instance.save()
                                if p_o_instance.process_status == '1':  
                                    p_o_instance.process_status = '2'  
                                    p_o_instance.save()  
                            
                            
                        messages.success(request, 'Purchase Order Quantities updated successfully.')
                        logger.info(f'Purchase Order Quantities updated-{form.instance.id}')

                        
                        return redirect('purchase-order-raw-material-list')

                    except DatabaseError as db_err:

                        logger.error(f'Database error during form save: {db_err}')
                        messages.error(request, 'A database error occurred during form save.')

                    except Exception as e:
                        logger.error(f'Unexpected error during form save: {e}')
                        messages.error(request, 'An unexpected error occurred during form save.')
                else:
                    
                    print(formset.errors)
                    logger.error(f'Purchase Order Quantities update error - {formset.errors} {formset.non_form_errors()}')
                    messages.error(request, f'There were errors in the form. Please correct them and try again.')

            except ValidationError as val_err:
                logger.error(f'Validation error: {val_err}-{formset.errors}')
                messages.error(request, f'Validation error: {val_err}')

            except Exception as e:
                logger.error(f'Exception occurred: {e}')
                messages.error(request, f'An exception occurred: {e}') 
        

    return render(request,'production/purchaseordercreateupdate.html',{'form':form ,'formset':formset,'raw_material_godowns':raw_material_godowns,
                                                                          'ledger_party_names':ledger_party_names,
                                                                          "products":products,'model_name':model_name,'model_images':model_images})



@login_required(login_url='login')
def purchaseorderlist(request):
    purchase_orders = purchase_order.objects.all().order_by('created_date')
    return render(request,'production/purchaseorderlist.html',{'purchase_orders': purchase_orders})



@login_required(login_url='login')
def purchaseorderdelete(request,pk):

    try:
        instance = get_object_or_404(purchase_order, pk = pk)
        instance.delete()
        logger.info(f"Purchase Order with order no - {instance.purchase_order_number} was deleted")
        messages.success(request,f'Purchase order with order no - {instance.purchase_order_number} was deleted')
        
    except Exception as e:
        messages.error(request,f'Cannot delete {instance.purchase_order_number} - {e}.')
        logger.error(f"Cannot delete {instance.purchase_order_number} - {e}.")
    return redirect('purchase-order-list')
     

def excel_download_production(request, module_name, pk):


    if module_name is not None and pk is not None:  

        file_name = None
        
        
        thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"))
        
        if module_name == 'purchase_order_raw':
            wb = Workbook()

            
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)  
            wb.create_sheet('purchase_order_raw')

            sheet = wb.worksheets[0]

            file_name = 'purchase_order_raw'

            column_widths = [13, 16, 10, 30, 13, 10, 10, 15, 15]  

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            purchase_order_instance = purchase_order.objects.get(pk=pk)
            
            sheet.cell(row=1, column=1).value = f'PO No:  {purchase_order_instance.purchase_order_number}'
            
            sheet.cell(row=1, column=1).font = Font(bold=True)
            
            sheet.cell(row=1, column=1).border = thin_border
            sheet.cell(row=1, column=2).border = thin_border
            sheet.cell(row=1, column=3).border = thin_border
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=2, column=1).value = f'Date:  {purchase_order_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'
            
            sheet.cell(row=2, column=1).font = Font(bold=True)

            sheet.cell(row=2, column=1).border = thin_border
            sheet.cell(row=2, column=2).border = thin_border
            sheet.cell(row=2, column=3).border = thin_border
            sheet.cell(row=2, column=4).border = thin_border

            sheet.cell(row=3, column=1).value = f'Ref No: { purchase_order_instance.product_reference_number.Product_Refrence_ID}'
            
            sheet.cell(row=3, column=1).font = Font(bold=True)

            sheet.cell(row=3, column=1).border = thin_border
            sheet.cell(row=3, column=2).border = thin_border
            sheet.cell(row=3, column=3).border = thin_border
            sheet.cell(row=3, column=4).border = thin_border

            sheet.cell(row=4, column=1).value = f'Name:  {purchase_order_instance.product_reference_number.Model_Name}'
            
            sheet.cell(row=4, column=1).font = Font(bold=True)
           
            sheet.cell(row=4, column=1).border = thin_border
            sheet.cell(row=4, column=2).border = thin_border
            sheet.cell(row=4, column=3).border = thin_border
            sheet.cell(row=4, column=4).border = thin_border

            sheet.cell(row=5, column=1).value = f'Party Name:  {purchase_order_instance.ledger_party_name.name}'
            
            sheet.cell(row=5, column=1).font = Font(bold=True)

            sheet.cell(row=5, column=1).border = thin_border
            sheet.cell(row=5, column=2).border = thin_border
            sheet.cell(row=5, column=3).border = thin_border
            sheet.cell(row=5, column=4).border = thin_border

            sheet.cell(row=6, column=1).value = f'Total PO Qty: {purchase_order_instance.number_of_pieces}'
           
            sheet.cell(row=6, column=1).font = Font(bold=True)

            sheet.cell(row=6, column=1).border = thin_border
            sheet.cell(row=6, column=2).border = thin_border
            sheet.cell(row=6, column=3).border = thin_border
            sheet.cell(row=6, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Product SKU'
            sheet.cell(row=1, column=6).value = 'Color'
            sheet.cell(row=1, column=8).value = 'Proc Qty'
            
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=8).font = Font(bold=True)
            
            sheet.cell(row=1, column=3).border = thin_border
            sheet.cell(row=1, column=4).border = thin_border
            sheet.cell(row=1, column=5).border = thin_border
            sheet.cell(row=1, column=6).border = thin_border
            sheet.cell(row=1, column=7).border = thin_border
            

            
            start_row = 2
            start_column = 5

            product_2_item_qs = purchase_order_instance.p_o_to_products.all()

            for index, instance in enumerate(product_2_item_qs.order_by('id'), start = start_row): 

                sheet.cell(row=index, column=start_column).value = instance.product_id.PProduct_SKU
                sheet.cell(row=index, column=start_column).border = thin_border
                sheet.cell(row=index, column=start_column + 1).value = instance.product_id.PProduct_color.color_name
                sheet.cell(row=index, column=start_column + 1).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 1).border = thin_border
                sheet.cell(row=index, column=start_column + 2).border = thin_border
                sheet.cell(row=index, column=start_column + 3).value = instance.order_quantity
                sheet.cell(row=index, column=start_column + 3).border = thin_border
                
                product_img_instance = instance.product_id.PProduct_image

                if product_img_instance:

                    product_img = product_img_instance.url

                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))

                
                
                
            
                
                if os.path.isfile(full_image_path):

                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    

                    
                    excel_img.width = 55  
                    excel_img.height = 50  

                    
                    

                    img_position = f'I{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:
                    sheet.cell(row=index, column=start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

            length_queryset = len(product_2_item_qs)
            
            row_num = 2 + length_queryset

            sheet.cell(row=row_num, column = 6).value = 'Total'
            sheet.cell(row=row_num, column = 8).value = purchase_order_instance.number_of_pieces
            sheet.cell(row=row_num, column = 5).font = Font(bold=True)
            sheet.cell(row=row_num, column = 8).font = Font(bold=True)
            sheet.cell(row=row_num, column = 6).border = thin_border
            sheet.cell(row=row_num, column = 8).border = thin_border
            sheet.cell(row=row_num, column = 7).border = thin_border
            




            
            if length_queryset < 6:

                start_row_items = 9

            else:
                start_row_items = length_queryset + 4

            start_column_items = 1 

            header_row = start_row_items - 1
            
            headers = ["Body/Combi", "Product Color", "Pcs", "Material Name", "Rate", "Panha","Units", "Consmp","Combi-2-Consump","T-Consmp","Physical Stock","Bal Stock"]

            
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)

            for index, instance in enumerate(purchase_order_instance.raw_materials.all().order_by('id'), start=start_row_items):
                sheet.cell(row=index, column = start_column_items).value = instance.Remark

                if instance.product_color != 'Common Item':
                    p_color = instance.product_color

                else:

                    p_color = ''

                sheet.cell(row=index, column=start_column_items + 1).value = p_color 
                sheet.cell(row=index, column=start_column_items + 2).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 3).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 4).value = instance.rate
                sheet.cell(row=index, column=start_column_items + 5).value = instance.panha
                sheet.cell(row=index, column=start_column_items + 6).value = instance.units
                sheet.cell(row=index, column=start_column_items + 7).value = instance.consumption
                sheet.cell(row=index, column=start_column_items + 8).value = instance.combi_consumption
                sheet.cell(row=index, column=start_column_items + 9).value = instance.total_comsumption
                sheet.cell(row=index, column=start_column_items + 10).value = instance.physical_stock
                sheet.cell(row=index, column=start_column_items + 11).value = instance.balance_physical_stock

            qs_length_list = len(purchase_order_instance.raw_materials.all())
            sheet.cell(row=qs_length_list + start_row_items, column=2).value =  'Narration'       
            sheet.cell(row=qs_length_list + start_row_items, column=3).value =  purchase_order_instance.note       
            sheet.cell(row=qs_length_list + start_row_items, column=2).font = Font(bold = True)
            sheet.cell(row=qs_length_list + start_row_items, column=3).font = Font(bold = True)




        elif module_name == 'purchase_order_cutting':
            wb = Workbook()

            
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)  
            wb.create_sheet('Cutting_Order')
            wb.create_sheet('Cutting_size')

            sheet = wb.worksheets[0]
            sheet1 = wb.worksheets[1]

            file_name = 'purchase_order_cutting'

            column_widths = [16, 20, 15, 15, 15, 15, 15, 15, 15]  

            column_widths_sheet_2 = [30, 25, 20, 20, 20, 20, 12, 12, 12]

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            for i, column_width in enumerate(column_widths_sheet_2, start=1):  
                col_letter = get_column_letter(i)
                sheet1.column_dimensions[col_letter].width = column_width

        
            purchase_order_cutting_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id=pk)
            
            sheet.cell(row=1, column=1).value = f'P O No:  {purchase_order_cutting_instance.purchase_order_id.purchase_order_number}'

            sheet.cell(row=2, column=1).value = f'Cutting No:  {purchase_order_cutting_instance.raw_material_cutting_id}'
            
            sheet.cell(row=3, column=1).value = f'Date:  {purchase_order_cutting_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'  
            
            sheet.cell(row=4, column=1).value = f'Cutter Name:  {purchase_order_cutting_instance.factory_employee_id.factory_emp_name}'
            
            sheet.cell(row=5, column=1).value = f'Prod Ref No:  {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID}'

            sheet.cell(row=6, column=1).value = f'Model Name:  {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Model_Name}'
        
            sheet.cell(row=7, column=1).value = f'Processed Qty:  {purchase_order_cutting_instance.processed_qty}'
            
            sheet.cell(row=9, column=1).value = f'Party Name:  {purchase_order_cutting_instance.purchase_order_id.ledger_party_name.name}'
           
            sheet.cell(row=10, column=1).value = f'Total PO Qty:  {purchase_order_cutting_instance.purchase_order_id.number_of_pieces}'
            
            sheet.cell(row=11, column=1).value = f'Bal Qty:  {purchase_order_cutting_instance.balance_qty}'
            
            sheet.cell(row=12, column=1).value = f'Target Date:  {purchase_order_cutting_instance.purchase_order_id.target_date.strftime("%d %B %Y")}'



            for index in range(1,13):
            
                sheet.cell(row=index, column=1).font = Font(bold=True)
            
                sheet.cell(row=index, column=1).border = thin_border
                sheet.cell(row=index, column=2).border = thin_border
                sheet.cell(row=index, column=3).border = thin_border

                
            sheet.cell(row=1, column=4).value = 'Product SKU'
            sheet.cell(row=1, column=4).font = Font(bold=True)
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Color'
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=5).border = thin_border

            sheet.cell(row=1, column=6).value = 'Cutting Qty'
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=6).border = thin_border

            
            
            
            start_row = 2
            start_column = 4  

            for index, instance in enumerate(purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all().order_by('id'), start=start_row):

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:
                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    
                    img_position = f'G{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:

                    sheet.cell(row=index, column=start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    


                sheet.cell(row=index, column=start_column).value = instance.product_sku
                sheet.cell(row=index, column=start_column).font = Font(bold=True)
                sheet.cell(row=index, column=start_column).border = thin_border


                sheet.cell(row=index, column=start_column + 1).value = instance.product_color
                sheet.cell(row=index, column=start_column + 1).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 1).border = thin_border


                sheet.cell(row=index, column=start_column + 2).value = instance.cutting_quantity
                sheet.cell(row=index, column=start_column + 2).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 2).border = thin_border



            qs_length = len(purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all())

            sheet.cell(row=qs_length + 2, column = 5).value = 'Total'
            sheet.cell(row=qs_length + 2, column = 5).border = thin_border
            sheet.cell(row=qs_length + 2, column = 5).font = Font(bold=True)
            sheet.cell(row=qs_length + 2, column = 6).value = purchase_order_cutting_instance.processed_qty
            sheet.cell(row=qs_length + 2, column = 6).border = thin_border
            sheet.cell(row=qs_length + 2, column = 6).font = Font(bold=True)


            if  qs_length < 11:

                
                start_row_items = 14

            else:

                start_row_items = qs_length + 4

            start_column_items = 1 

            header_row = start_row_items - 1

            
            headers = ["Body/Combi", "Product Color","Pcs","Material Name", 'Shade Color',"Panha","Units","Consump","Combi-2-Consump","Total Consump","Physical Stock","Balance Stock"]

            
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)


            cutting_items_qs = purchase_order_cutting_instance.purchase_order_for_raw_material_cutting_items_set.filter(material_color_shade__items__Fabric_nonfabric='Fabric')

            for index, instance in enumerate(cutting_items_qs.order_by('id'), start=start_row_items):
                sheet.cell(row=index, column=start_column_items).value = instance.Remark


                if instance.product_color != 'Common Item':
                    p_color = instance.product_color

                else:
                    p_color = ''

                sheet.cell(row=index, column=start_column_items + 1).value = p_color
                sheet.cell(row=index, column=start_column_items + 2).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 3).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 4).value = instance.material_color_shade.item_shade_name
                
                sheet.cell(row=index, column=start_column_items + 5).value = instance.panha
                sheet.cell(row=index, column=start_column_items + 6).value = instance.units
                sheet.cell(row=index, column=start_column_items + 7).value = instance.consumption
                sheet.cell(row=index, column=start_column_items + 8).value = instance.combi_consumption
                sheet.cell(row=index, column=start_column_items + 9).value = instance.total_comsumption
                sheet.cell(row=index, column=start_column_items + 10).value = instance.physical_stock
                sheet.cell(row=index, column=start_column_items + 11).value = instance.balance_physical_stock


            qs_length_list = len(cutting_items_qs)

            sheet.cell(row=qs_length_list + start_row_items, column=2).value =  'Narration :'       
            sheet.cell(row=qs_length_list + start_row_items, column=3).value =  purchase_order_cutting_instance.note
            sheet.cell(row=qs_length_list + start_row_items, column=2).font = Font(bold=True)
            sheet.cell(row=qs_length_list + start_row_items, column=3).font = Font(bold=True)


            

            sheet1.cell(row=1, column=1).value = f'Ref No - {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID}'
            sheet1.cell(row=2, column=1).value = f'Model Name - {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Model_Name}'

            sheet1.cell(row=1, column=1).font = Font(bold=True)
            sheet1.cell(row=1, column=1).border = thin_border

            sheet1.cell(row=2, column=1).font = Font(bold=True)
            sheet1.cell(row=2, column=1).border = thin_border
            
            sheet1.cell(row = 1, column=2).value = 'Product SKU'
            sheet1.cell(row = 1, column=2).font = Font(bold=True)
            sheet1.cell(row = 1, column=2).border = thin_border

            sheet1.cell(row = 1, column=3).value = 'color'
            sheet1.cell(row = 1, column=3).font = Font(bold=True)
            sheet1.cell(row = 1, column=3).border = thin_border

            sheet1.cell(row = 1, column=4).font = Font(bold=True)
            sheet1.cell(row = 1, column=4).border = thin_border

             
            start_row = 2
            start_column = 2

            purchase_order_cutting_p_2_item_qs = purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all()

            for index, instance in enumerate(purchase_order_cutting_p_2_item_qs.order_by('id'), start=start_row):

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:
                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                    
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    
                    img_position = f'D{index}'

                    
                    sheet1.add_image(excel_img, img_position)

                else:

                    sheet1.cell(row=index, column=start_column + 2).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

                sheet1.cell(row=index, column=start_column).value = instance.product_sku
                sheet1.cell(row=index, column=start_column).font = Font(bold=True)
                sheet1.cell(row=index, column=start_column).border = thin_border

                sheet1.cell(row=index, column=start_column + 1).value = instance.product_color
                sheet1.cell(row=index, column=start_column + 1).font = Font(bold = True)
                sheet1.cell(row=index, column=start_column + 1).border = thin_border

                
                product_ref_no = purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID

                items_qs = product_2_item_through_table.objects.filter(
                    PProduct_pk__Product__Product_Refrence_ID = product_ref_no , Remark = 'BODY')


                
                for index, instance in enumerate(items_qs, start = 1):
                    sheet1.cell(row=index , column=5).value = instance.Item_pk.item_name
                    sheet1.cell(row=index , column=5).font = Font(bold=True)

            
            if len(purchase_order_cutting_p_2_item_qs) < len(items_qs) :
                start_row_items = len(items_qs) + 2
            
            else:
                start_row_items = len(purchase_order_cutting_p_2_item_qs) + 3
                
            start_column_items = 4 

            header_row = start_row_items - 1

            
            headers = ["Item Name", "Part Name", "Part Dimentions", 'Dimention Total','Part Pcs','Body/Combi','Grand Total','Grand Total Combi']

            
            for col_num, header in enumerate(headers, start = 1):
                sheet1.cell(row = header_row, column = col_num).value = header
                sheet1.cell(row = header_row, column = col_num).font = Font(bold=True)
                sheet1.cell(row = header_row, column = col_num).alignment = Alignment(wrap_text = True)

            start_column_items = 1

            product_with_combi = [] 

            combi_found = False

            
            for product in product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID = product_ref_no):

                if product.product_item_configs.filter(body_combi='combi').exists():
                    product_with_combi.append(product)
                    combi_found = True
            
            
            if not combi_found:

                body_instance = product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID = product_ref_no).first()

                product_with_combi.append(body_instance)
                

            for instance in product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU = product_with_combi[0].PProduct_pk.PProduct_SKU).order_by('row_number'):
                list_to_append = []
                for record in instance.product_item_configs.all().order_by('id'):
                    list_1 = [
                        record.producttoitem.Item_pk.item_name,
                        record.part_name,
                        record.part_dimentions,
                        record.dimention_total,
                        record.part_pieces,
                        record.body_combi
                        ]
                    
                    list_to_append.append(list_1)

                for row in list_to_append:
                    sheet1.append(row)

                sheet1.append(['','','','','','', instance.grand_total, instance.grand_total_combi])
                    

        elif module_name == 'labour_workout':

            wb = Workbook()

            
            default_sheet = wb['Sheet']

            wb.remove(default_sheet)  

            wb.create_sheet('Labour_workout')
            
            sheet = wb.worksheets[0]

            file_name = 'labour_workout'

            
            column_widths = [35, 20, 12, 10, 15, 30, 15, 15, 15]  

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            labour_workout_instance = labour_workout_childs.objects.get(id=pk)

            sheet.cell(row=1, column=1).value = f'PO No. : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number}'
            
            sheet.cell(row=2, column=1).value = f'Challan No : {labour_workout_instance.challan_no}, Date : {labour_workout_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'
            
            sheet.cell(row=3, column=1).value = f'Vendor Name : {labour_workout_instance.labour_name.name}'
            
            sheet.cell(row=4, column=1).value = f'Ref No : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID}, Model Name : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name}'
            
            sheet.cell(row=5, column=1).value = f'No of Pcs : {labour_workout_instance.total_process_pcs}'
            

            for cell in range(1,6):
                sheet.cell(row=cell, column=1).border = thin_border
                sheet.cell(row=cell, column=2).border = thin_border
                sheet.cell(row=cell, column=3).border = thin_border

            sheet.cell(row=1, column=4).value = 'Product SKU'
            sheet.cell(row=1, column=4).font = Font(bold=True)
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Color'
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=5).border = thin_border

            sheet.cell(row=1, column=6).value = 'Process Qty'
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=6).border = thin_border

            
            start_row = 2 
            start_column = 4

            labour_workout_p2i_items_qs = labour_workout_instance.labour_workout_child_items.all()

            for index, instance in enumerate(labour_workout_p2i_items_qs.order_by('id'), start = start_row):
                sheet.cell(row=index, column=start_column).value = instance.product_sku
                
                sheet.cell(row=index, column=start_column).border = thin_border

                sheet.cell(row=index, column=start_column + 1).value = instance.product_color
                
                sheet.cell(row=index, column=start_column + 1).border = thin_border

                sheet.cell(row=index, column=start_column + 2).value = instance.processed_pcs
                
                sheet.cell(row=index, column=start_column + 2).border = thin_border

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:

                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 

                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    col_letter =  get_column_letter(index) 
                    img_position = f'G{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:
                    sheet.cell(row=index, column = start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

            qs_length = len(labour_workout_p2i_items_qs)

            sheet.cell(row=qs_length + 2, column = 4).value = 'Total'
            sheet.cell(row=qs_length + 2, column = 4).border = thin_border
            sheet.cell(row=qs_length + 2, column = 4).font = Font(bold=True)

            sheet.cell(row=qs_length + 2, column = 5).value = labour_workout_instance.total_process_pcs
            sheet.cell(row=qs_length + 2, column = 5).border = thin_border
            sheet.cell(row=qs_length + 2, column = 5).font = Font(bold=True)

            
            for index, instance in enumerate(product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.
                product_reference_number.Product_Refrence_ID, Remark='BODY'), start = 7):
                
                sheet.cell(row=index , column=1).value = instance.Item_pk.item_name
                
            len_items = len(product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.
                product_reference_number.Product_Refrence_ID, Remark='BODY'))

            
            if qs_length < len_items + 7:

                start_row_items = len_items + 8
            else:
                start_row_items = qs_length + 4
                
            header_row = start_row_items - 1

            
            
            headers = ["Item Name", "Part Name", "Total Part Pcs"]

            
            for col_num, header in enumerate(headers, start = 1):
                sheet.cell(row = header_row, column = col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text = True)

            product_with_combi = [] 

            combi_found = False

            
            for product in product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.
                purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID):
                
                if product.product_item_configs.filter(body_combi = 'combi').exists():
                    product_with_combi.append(product)
                    combi_found = True

            
            if not combi_found:
                body_instance = product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.
                purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID).first()

                product_with_combi.append(body_instance)

            rows_iterated_outer = 0

            for record in product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU = product_with_combi[0].PProduct_pk.PProduct_SKU,Item_pk__Fabric_nonfabric = 'Fabric').order_by('id'):
                list_to_append = []    

                rows_iterated_inner = 0
                item_name_last = ''
                for instance in record.product_item_configs.all().order_by('id'):

                    if instance.part_pieces is None or instance.part_pieces == 0 :

                        part_pcs = 0 
                    else:
                        part_pcs =  instance.part_pieces

                    if item_name_last == record.Item_pk.item_name:
                        item_name_present = ''

                    else:
                        item_name_present = record.Item_pk.item_name


                    list = [
                        item_name_present,
                        instance.part_name,
                        part_pcs * labour_workout_instance.total_process_pcs
                        
                        ]
                    rows_iterated_inner = rows_iterated_inner + 1

                    item_name_last = record.Item_pk.item_name

                    list_to_append.append(list)

                for x in list_to_append:
                    sheet.append(x)
                rows_iterated_outer = rows_iterated_outer + rows_iterated_inner + 1
                sheet.append(['', '' , ''])
            
            start_column_items = 4 

            
            headers = ["Body/Combi", "Pcs", "Material Name", "Total Consump"]

            
            for col_num, header in enumerate(headers, start=4):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)


            for index, instance in enumerate(labour_workout_instance.labour_workout_cutting_items_set.all().order_by('id'), start = start_row_items):
                sheet.cell(row=index, column=start_column_items).value = instance.Remark
                sheet.cell(row=index, column=start_column_items + 1).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 2).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 3).value = instance.total_comsumption


            narration_row = rows_iterated_outer + len_items + 8
        
            sheet.cell(row=narration_row, column=2).value = f'Narration: {labour_workout_instance.note}'
            sheet.cell(row=narration_row, column=2).font = Font(bold=True)

            for cell in range(1,narration_row):
                sheet.cell(row=cell, column=1).font = Font(bold=True)


        
        fileoutput = BytesIO()
        wb.save(fileoutput)
            
        
        response = HttpResponse(fileoutput.getvalue(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name_with_pk = f'product_reference_id_{file_name}'
        response['Content-Disposition'] = f'attachment; filename="{file_name_with_pk}.xlsx"'

        return response
    
    else:
        return HttpResponse('INVALID ENTRY')



@login_required(login_url = 'login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseorderrawmaterial(request ,p_o_pk, prod_ref_no):
    
    purchase_order_instance = purchase_order.objects.get(pk=p_o_pk)

    form = purchase_order_form(instance = purchase_order_instance)
    
    product_refrence_no = prod_ref_no

    product_2_items_instances_unique = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no, common_unique = False).order_by(
                                'row_number')
    
    product_2_items_instances_common = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no, common_unique = True).order_by(
                                'row_number','id').distinct('row_number')
    
    
    
    
    product_2_items_instances = product_2_items_instances_unique.union(product_2_items_instances_common)
    

    model_name = purchase_order_instance.product_reference_number.Model_Name

    physical_stock_all_godowns = {}

    for item in product_2_items_instances:
        item_id = item.Item_pk
        item_name = item.Item_pk.item_name
        item_quantity = 0
        item_godowns = item_godown_quantity_through_table.objects.filter(Item_shade_name__items = item_id,
                                        godown_name = purchase_order_instance.temp_godown_select) 
        
        if item_godowns:
            for query in item_godowns:
                item_quantity = item_quantity + query.quantity
                physical_stock_all_godowns[item_name] = str(item_quantity)

        else:
            physical_stock_all_godowns[item_name] = str(item_quantity)

    purchase_order_raw_formset = purchase_order_raw_product_qty_formset(instance = purchase_order_instance)



    
    if not purchase_order_instance.raw_materials.all():
        
        physical_stock_all_godown_json = json.dumps(physical_stock_all_godowns) 

        initial_data = []
        for query in product_2_items_instances:
            rate_first = query.Item_pk.shades.order_by('id').first() 
            
            
            if query.common_unique == True:
                product_color_or_common_item = 'Common Item'
                product_sku_or_common_item = 'Common Item'

            else:
                product_color_or_common_item = query.PProduct_pk.PProduct_color
                product_sku_or_common_item = query.PProduct_pk.PProduct_SKU
            
            initial_data_dict = {'product_sku': product_sku_or_common_item,
                                'product_color' : product_color_or_common_item,
                                'material_name':query.Item_pk.item_name,
                                'rate':rate_first.rate,
                                'panha':query.Item_pk.Panha,
                                'units':query.Item_pk.Units,
                                'g_total':query.grand_total,
                                'g_total_combi':query.grand_total_combi,
                                'consumption':'0',
                                'combi_consumption':0,
                                'total_comsumption':'0',
                                'unit_value': query.Item_pk.unit_name_item.unit_name,
                                'physical_stock':'0',
                                'balance_physical_stock':'0',
                                'row_number':query.row_number,
                                'Remark':query.Remark, 
                                'pcs': '0' }
            
            initial_data.append(initial_data_dict)

        initial_sorted_data = sorted(initial_data, key=itemgetter('row_number'), reverse=False)

        
        purchase_order_raw_product_sheet_formset = inlineformset_factory(purchase_order, purchase_order_for_raw_material, form=purchase_order_raw_product_sheet_form, extra=len(initial_sorted_data) if initial_data else 0, can_delete=False)

        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(initial=initial_sorted_data, instance=purchase_order_instance)


    
    elif purchase_order_instance.raw_materials.all():

        physical_stock_all_godown_json = None 

        purchase_order_raw_product_sheet_formset = inlineformset_factory(purchase_order, purchase_order_for_raw_material, form=purchase_order_raw_product_sheet_form, extra=0, can_delete=False)

        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(instance=purchase_order_instance)


    if request.method == 'POST':
        print(request.POST)
        purchase_order_raw_formset = purchase_order_raw_product_qty_formset(request.POST)

        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(request.POST, instance=purchase_order_instance)

        try:
            if purchase_order_raw_formset.is_valid() and purchase_order_raw_sheet_formset.is_valid():
                
                try:
                    with transaction.atomic():
                        
                        purchase_order_raw_formset.save()
                        purchase_order_raw_sheet_formset.save()
                        
                        note_post = request.POST.get('note')

                        if note_post:
                            purchase_order_instance.note = note_post
                        else:
                            purchase_order_instance.note = None
                            
                        purchase_order_instance.save()


                        for form in purchase_order_raw_sheet_formset:
                            po_form_instance = form.instance.purchase_order_id  
                            if po_form_instance.process_status == '2':   
                                po_form_instance.process_status = '3'  
                                po_form_instance.save()  

                        
                        return redirect('purchase-order-raw-material-list')
                
                except ValueError as ve:
                    messages.error(request,f'Error Occured - {ve}')

                except Exception as e:
                    messages.error(request,f'Exception Occured - {e}')
                
                return render(request,'production/purchaseorderrawmaterial.html',{'form': form ,'model_name':model_name,
                                                                        'purchase_order_raw_formset':purchase_order_raw_formset,
                                                                        'purchase_order_raw_sheet_formset':purchase_order_raw_sheet_formset,
                                                                        'physical_stock_all_godown_json':physical_stock_all_godown_json})

            
            else:
                print(purchase_order_raw_formset.errors,purchase_order_raw_sheet_formset.errors)
                raise ValidationError("No products found for the given reference ID.")
                
                
                
                
                
        except ValidationError as ve:
                messages.error(request,f' Please enter correct Procurement color wise QTY {ve}')
        
        except Exception as e:
            messages.error(request,f' An exception occoured {e}')



    return render(request,'production/purchaseorderrawmaterial.html',{'form':form, 'model_name':model_name,
                                                                      'purchase_order_raw_formset':purchase_order_raw_formset,
                                                                      'purchase_order_raw_sheet_formset':purchase_order_raw_sheet_formset,
                                                                      'physical_stock_all_godown_json':physical_stock_all_godown_json})



@login_required(login_url='login')
def purchase_order_for_raw_material_list(request):
    
    
    purchase_orders_pending = purchase_order.objects.annotate(raw_material_count=Count('raw_materials')).filter(raw_material_count__lt=1, purchase_order_to_product_saved=True).order_by('created_date')
    purchase_orders_completed = purchase_order.objects.annotate(raw_material_count=Count('raw_materials')).filter(raw_material_count__gt=0).order_by('created_date')


    return render(request,'production/purchase_order_for_raw_material_list.html',
                  {'purchase_orders_pending': purchase_orders_pending,
                   'purchase_orders_completed':purchase_orders_completed})



@login_required(login_url='login')
def purchase_order_for_raw_material_delete(request,pk):

    try:
        purchase_order_raw_instances = purchase_order_for_raw_material.objects.filter(purchase_order_id=pk)

        for instances in purchase_order_raw_instances:
            instances.delete()
        messages.success(request,'Deleted Successfully')
        return redirect('purchase-order-raw-material-list')
    
    except Exception as e:
        messages.error(request,f'Error While Deleting purchase Order {e}')
        return redirect('purchase-order-raw-material-list')

 

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseordercuttingcreateupdate(request,p_o_pk,prod_ref_no,pk=None):

    if pk:
        purchase_order_cutting_instance = get_object_or_404(purchase_order_raw_material_cutting, pk=pk)
    else:
        purchase_order_cutting_instance = None

    labour_all = factory_employee.objects.all()

    
    purchase_order_instance = get_object_or_404(purchase_order, pk=p_o_pk)

    
    purchase_order_raw_instances = purchase_order_for_raw_material.objects.filter(purchase_order_id=p_o_pk).order_by('id')

    
    purchase_order_to_product_instances = purchase_order_to_product.objects.filter(purchase_order_id = p_o_pk)

    
    form = purchase_order_form(instance = purchase_order_instance)

    current_godown = form.instance.temp_godown_select

    
    purchase_order_cutting_form = purchase_order_raw_material_cutting_form(request.POST or None, instance=purchase_order_cutting_instance)

    
    if not pk:

        product_refrence_no = prod_ref_no

    
        product_2_items_instances_unique = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no,common_unique = False).order_by(
                                'row_number')
    
        product_2_items_instances_common = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no,common_unique = True).order_by(
                                'row_number','id').distinct('row_number')


        product_2_items_instances = product_2_items_instances_unique.union(product_2_items_instances_common)

        
        initial_data = []

        for purchase_items_raw in product_2_items_instances:
            
            current_godown_qty = item_color_shade.objects.filter(items=purchase_items_raw.Item_pk,godown_shades__godown_name=purchase_order_instance.temp_godown_select).annotate(total_godown_qty = Sum('godown_shades__quantity'))

            rate_first = purchase_items_raw.Item_pk.shades.order_by('id').first() 

            shade_single_list = []  
            
            
            for qs in purchase_items_raw.Item_pk.shades.all():
                try:
                    
                    shade_godown_qty = item_godown_quantity_through_table.objects.get(
                    godown_name=purchase_order_instance.temp_godown_select,
                    Item_shade_name=qs.id)

                    
                    shade_data = {
                    'shade_id': qs.id,
                    'shade_name': qs.item_shade_name,
                    'godown_quantity': shade_godown_qty.quantity}  
                        
                    shade_single_list.append(shade_data)

                except item_godown_quantity_through_table.DoesNotExist:
                    

                    shade_data = {
                        'shade_id': qs.id,
                        'shade_name': qs.item_shade_name,
                        'godown_quantity': 0 }
                    
                    shade_single_list.append(shade_data)


            if purchase_items_raw.common_unique == True:
                product_color_or_common_item = 'Common Item'
                product_sku_or_common_item = 'Common Item'

            else:
                product_color_or_common_item = purchase_items_raw.PProduct_pk.PProduct_color
                product_sku_or_common_item = purchase_items_raw.PProduct_pk.PProduct_SKU

            current_godown_qty_total = 0

            for qs in current_godown_qty:
                current_godown_qty_total = current_godown_qty_total + qs.total_godown_qty

            
            initial_data_dict = {
                'product_sku': product_sku_or_common_item,
                'product_color' : product_color_or_common_item,
                'material_name' : purchase_items_raw.Item_pk.item_name,
                'material_color_shade': shade_single_list, 
                'fabric_non_fab': purchase_items_raw.Item_pk.Fabric_nonfabric, 
                'rate' : rate_first.rate,
                'panha' : purchase_items_raw.Item_pk.Panha,
                'units' :  purchase_items_raw.Item_pk.Units,
                'g_total' : purchase_items_raw.grand_total,
                'g_total_combi': purchase_items_raw.grand_total_combi,
                'consumption' : '0',
                'total_comsumption' :'0',
                'combi_consumption' : '0',
                'unit_value' : purchase_items_raw.Item_pk.unit_name_item.unit_name,
                'physical_stock' : current_godown_qty_total,
                'balance_physical_stock': '0',
                'row_number': purchase_items_raw.row_number,
                'Remark': purchase_items_raw.Remark,
                'pcs': '0'}
            
            initial_data.append(initial_data_dict)

       
        initial_sorted_data = sorted(initial_data, key = itemgetter('row_number'), reverse=False)

        
        
        purchase_order_for_raw_material_cutting_items_formset = inlineformset_factory(purchase_order_raw_material_cutting, 
                                                                                      purchase_order_for_raw_material_cutting_items, 
                                                                                      form=purchase_order_for_raw_material_cutting_items_form,
                                                                                      formset=Basepurchase_order_for_raw_material_cutting_items_form, 
                                                                                      extra=len(initial_data), 
                                                                                      can_delete=False)


        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(initial=initial_sorted_data)


        
        initial_data_p_o_to_items = []

        for instances in purchase_order_to_product_instances:
            initial_data_dict = {
                'product_color': instances.product_id.PProduct_color,
                'product_sku': instances.product_id.PProduct_SKU,
                'order_quantity': instances.order_quantity,
                'process_quantity': instances.process_quantity,
                'cutting_quantity': '0',
            }

            initial_data_p_o_to_items.append(initial_data_dict)

        
        purchase_order_to_product_formset = inlineformset_factory(purchase_order_raw_material_cutting, 
                                                                  purchase_order_to_product_cutting,
                                                                    form=purchase_order_to_product_cutting_form,
                                                                      extra=len(initial_data_p_o_to_items),
                                                                      can_delete=False)

        purchase_order_to_product_formset_form = purchase_order_to_product_formset(initial= initial_data_p_o_to_items)

    
    elif pk:

        purchase_order_for_raw_material_cutting_items_formset = inlineformset_factory(purchase_order_raw_material_cutting,
                                                                                       purchase_order_for_raw_material_cutting_items,
                                                                                         form=purchase_order_for_raw_material_cutting_items_form, 
                                                                                         extra=0, can_delete=False)
        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(instance=purchase_order_cutting_instance)


        
        purchase_order_to_product_formset = inlineformset_factory(purchase_order_raw_material_cutting, purchase_order_to_product_cutting, 
                                                                  form=purchase_order_to_product_cutting_form, extra=0, can_delete=False)
        
        purchase_order_to_product_formset_form = purchase_order_to_product_formset(instance=purchase_order_cutting_instance)
        
    if request.method == 'POST':
        
        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(request.POST)

        
        purchase_order_to_product_formset_form = purchase_order_to_product_formset(request.POST)

        if purchase_order_cutting_form.is_valid() and purchase_order_to_product_formset_form.is_valid() and purchase_order_for_raw_material_cutting_items_formset_form.is_valid():
            try:
                with transaction.atomic():
                    
                    cutting_form_instance = purchase_order_cutting_form.save()
                    cutting_form_instance.purchase_order_id.cutting_total_processed_qty = cutting_form_instance.purchase_order_id.cutting_total_processed_qty + cutting_form_instance.processed_qty
                    cutting_form_instance.purchase_order_id.save()
                    
                    
                    if cutting_form_instance.purchase_order_id.process_status == '3':
                        cutting_form_instance.purchase_order_id.process_status = '4'
                        cutting_form_instance.purchase_order_id.save()
                    

                    
                    for form in purchase_order_to_product_formset_form:

                        if form.is_valid(): 
                            try:
                                p_o_to_order_form_instance = form.save(commit = False)
                                p_o_to_order_form_instance.purchase_order_cutting_id = cutting_form_instance
                                p_o_to_order_form_instance.save()

                                
                                product_sku = p_o_to_order_form_instance.product_sku
                                processed_qty = p_o_to_order_form_instance.cutting_quantity

                                p_o_id = p_o_to_order_form_instance.purchase_order_cutting_id.purchase_order_id
                            
                                purchase_order_products = purchase_order_to_product.objects.filter(purchase_order_id =p_o_id,product_id =product_sku).first()

                                if purchase_order_products:
                                    purchase_order_products.process_quantity =  purchase_order_products.process_quantity - processed_qty
                                    purchase_order_products.save()

                                else:
                                    logger.error(f'Product {product_sku} not found in purchase order {p_o_id}')
                                    messages.error(request, f'Product {product_sku} not found in the purchase order.')

                            except Exception as e:
                                logger.error(f'Error processing product form: {e}')
                                messages.error(request, f'An error occurred while processing the product form: {e}')


                    
                    for form in purchase_order_for_raw_material_cutting_items_formset_form:
                        if form.is_valid(): 
                            try:
                                form_instance = form.save(commit=False)
                                form_instance.purchase_order_cutting = cutting_form_instance
                                form_instance.total_comsumption_in_cutting = form_instance.total_comsumption
                                form_instance.entry_from_cutting_room = True
                                form_instance.save()
                            

                            except ValidationError as ve:
                                logger.error(f'Error with: {e}')

                            except Exception as e:
                                logger.error(f'Error processing raw material form: {e}')
                                messages.error(request, f'An error occurred while processing the raw material form: {e}')
                        
                        
                    
                    processed_quantity = int(request.POST['processed_qty'])
                    qty_to_process = cutting_form_instance.purchase_order_id.balance_number_of_pieces  
                    qty_to_process_minus_processed_qty = qty_to_process - processed_quantity  
                    cutting_form_instance.purchase_order_id.balance_number_of_pieces = qty_to_process_minus_processed_qty  
                    cutting_form_instance.purchase_order_id.save() 

                    messages.success(request, f'Cutting Order Created SuccessFully')
                    return(redirect(reverse('purchase-order-cutting-list', args = [cutting_form_instance.purchase_order_id.id, cutting_form_instance.purchase_order_id.product_reference_number.Product_Refrence_ID])))


            except ValidationError as val_err:
                logger.error(f'Validation error: {val_err}')
                messages.error(request, f'Validation error: {val_err}')


            except DatabaseError as db_err:
                logger.error(f'Database error: {db_err}')
                messages.error(request, f'Database error: {db_err}')


            except Exception as e:
                logger.error(f'Unexpected error: {e}')
                messages.error(request, f'An unexpected error occurred: {e}')

        else:
            
            
            

            
            


            return render(request,'production/purchase_order_cutting.html',{'form':form,'labour_all':labour_all,'purchase_order_cutting_form':purchase_order_cutting_form,'p_o_pk':p_o_pk,
                                                                    'purchase_order_to_product_formset_form':purchase_order_to_product_formset_form,
                                                                     'purchase_order_for_raw_material_cutting_items_formset_form':purchase_order_for_raw_material_cutting_items_formset_form})


    return render(request,'production/purchase_order_cutting.html',{'form':form,'labour_all':labour_all,'purchase_order_cutting_form':purchase_order_cutting_form,'p_o_pk':p_o_pk,
                                                                    'purchase_order_to_product_formset_form':purchase_order_to_product_formset_form,
                                                                     'purchase_order_for_raw_material_cutting_items_formset_form':purchase_order_for_raw_material_cutting_items_formset_form})




@login_required(login_url='login')
def purchaseordercuttinglistall(request):

    
    raw_materials_exists = purchase_order_for_raw_material.objects.filter(purchase_order_id=OuterRef('pk')
                                                ).values('pk')[:1] 

    
    purchase_orders_cutting_pending = (
        purchase_order.objects.annotate(
            raw_materials_exist=Exists(raw_materials_exists),
            total_approved_qty_sum=Sum('cutting_pos__approved_qty'),
        )
        .filter(
            balance_number_of_pieces__gt=0,
            raw_materials_exist=True  
        )
        .annotate(
            total_approved_balance=F('number_of_pieces') - F('total_approved_qty_sum')
        )
        .order_by('created_date')
    )

    purchase_orders_cutting_completed = purchase_order.objects.filter(
        balance_number_of_pieces=0).annotate(total_processed_qty = Sum('cutting_pos__processed_qty'),
        total_approved_qty_sum = Sum('cutting_pos__approved_qty')).annotate(total_approved_balance = F(
        'number_of_pieces') - F('total_approved_qty_sum')).order_by('created_date')


    return render(request,'production/purchaseordercuttinglistall.html', {'purchase_orders_cutting_pending':purchase_orders_cutting_pending,'purchase_orders_cutting_completed':purchase_orders_cutting_completed})



@login_required(login_url='login')
def purchaseordercuttinglist(request,p_o_pk,prod_ref_no):
    p_o_cutting_order_all = purchase_order_raw_material_cutting.objects.filter(
        purchase_order_id = p_o_pk).select_related('purchase_order_id__ledger_party_name',
        'factory_employee_id').order_by('created_date')


    Purchase_order_no = purchase_order.objects.get(id=p_o_pk)
    return render(request,'production/purchaseordercuttinglist.html', {'p_o_cutting_order_all':p_o_cutting_order_all, 'p_o_number':Purchase_order_no, 'prod_ref_no':prod_ref_no, 'p_o_pk':p_o_pk})

@login_required(login_url='login')
def purchaseordercuttingapprovalcheckajax(request):
    cutting_pk_no = request.GET.get('cutting_pk_no')

    if not cutting_pk_no:
        return JsonResponse({'status': 'error', 'message': 'Cutting ID not provided'}, status=400)

    try:
        
        purchase_order_master_instances = labour_workout_master.objects.filter(
            purchase_order_cutting_master__raw_material_cutting_id=cutting_pk_no
        ).order_by('created_date')

        if not purchase_order_master_instances.exists():
            return JsonResponse({'status': 'error', 'message': 'No records found for the given Cutting ID'}, status=404)

        approval_data = []
        for x in purchase_order_master_instances:
            dict_to_append = {
                'Approved_Date': x.created_date,
                'Approved_Name': 'approved name',  
                'Approved_Qty': x.total_approved_pcs,
            }
            approval_data.append(dict_to_append)

        return JsonResponse({'status': 'success','approval_data': approval_data})

    except ObjectDoesNotExist as e:
        logger.error(f'Record not found: {e}')
        return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)

    except ValidationError as e:
        logger.error(f'Invalid data: {e}')
        return JsonResponse({'status': 'error', 'message': 'Invalid input data'}, status=400)

    except Exception as e:
        logger.error(f'Unexpected error: {e}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred'}, status = 500)



@login_required(login_url='login')
def pendingapprovall(request):
    pending_approval_query = purchase_order_raw_material_cutting.objects.exclude(processed_qty = F('approved_qty')).order_by('created_date') 
    return render(request,'production/cuttingapprovallistall.html', {'pending_approval_query': pending_approval_query})


@login_required(login_url = 'login')
def purchaseordercuttingpopup(request, cutting_id):

    if cutting_id:
        cutting_order_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id = cutting_id)
    
    else:
        cutting_order_instance = None

    formset = purchase_order_cutting_approval_formset(request.POST or None, instance = cutting_order_instance)

    if request.method == 'POST':
        if formset.is_valid():
            
            if any(form.has_changed() for form in formset): 
                formset_instance = formset.save(commit=False)

                
                raw_material_cutting_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id=cutting_id)

                
                old_total_approved_qty_total = raw_material_cutting_instance.approved_qty

                
                labour_workout_master_instance = labour_workout_master.objects.create(purchase_order_cutting_master=raw_material_cutting_instance)

                total_approved_pcs = 0

                for form in formset_instance:
                    p_o_to_cutting_instance = purchase_order_to_product_cutting.objects.get(id = form.id) 
                    old_total_approved_qty_diffrence  =  form.approved_pcs - p_o_to_cutting_instance.approved_pcs  
                    form.approved_pcs_diffrence = old_total_approved_qty_diffrence  
                    old_total_approved_qty_total = old_total_approved_qty_total + old_total_approved_qty_diffrence 
                    form.save() 
                    
                    total_approved_pcs = total_approved_pcs + old_total_approved_qty_diffrence

                    
                    product_to_item_labour_workout.objects.create(labour_workout = labour_workout_master_instance,
                                                                product_color=form.product_color,product_sku=form.product_sku,
                                                                pending_pcs = old_total_approved_qty_diffrence, processed_pcs = old_total_approved_qty_diffrence)

                raw_material_cutting_instance.approved_qty = old_total_approved_qty_total 
                raw_material_cutting_instance.save() 

                labour_workout_master_instance.total_approved_pcs = total_approved_pcs 
                labour_workout_master_instance.total_pending_pcs = total_approved_pcs 
                labour_workout_master_instance.save() 

            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """

            return HttpResponse(close_window_script)
        else:
            return render(request,'production/purchaseordercuttingpopup.html', {'formset':formset})
            
    return render(request,'production/purchaseordercuttingpopup.html', {'formset':formset})




@login_required(login_url='login')
def purchaseordercuttingmastercancelajax(request):

    if request.method == 'POST':

        try:
            cutting_key = request.POST.get('cuttingId')
            
            cutting_instance = get_object_or_404(purchase_order_raw_material_cutting,pk=cutting_key)

            with transaction.atomic():
                if cutting_instance:
                    cutting_instance.cutting_cancelled = True
                    cutting_instance.save()

                    if cutting_instance.approved_qty == 0:
                        processed_qty_to_revert = cutting_instance.processed_qty

                        
                        cutting_instance.purchase_order_id.cutting_total_processed_qty = cutting_instance.purchase_order_id.cutting_total_processed_qty - processed_qty_to_revert
                        cutting_instance.purchase_order_id.balance_number_of_pieces = cutting_instance.purchase_order_id.balance_number_of_pieces + processed_qty_to_revert
                        cutting_instance.purchase_order_id.save()
                        
                        
                        for record in cutting_instance.purchase_order_to_product_cutting_set.all():
                            purchase_order_to_product_instance = purchase_order_to_product.objects.get(purchase_order_id=cutting_instance.purchase_order_id.id,product_id__PProduct_SKU=record.product_sku)
                            purchase_order_to_product_instance.process_quantity = purchase_order_to_product_instance.process_quantity + record.cutting_quantity 
                            purchase_order_to_product_instance.save()

                        
                        for cutting_items in cutting_instance.purchase_order_for_raw_material_cutting_items_set.all():

                            item_godown_instance, created = item_godown_quantity_through_table.objects.get_or_create(Item_shade_name=cutting_items.material_color_shade,godown_name=cutting_items.purchase_order_cutting.purchase_order_id.temp_godown_select)

                            if created:
                                item_godown_instance_qty = 0 
                            else:
                                item_godown_instance_qty = item_godown_instance.quantity

                            item_godown_instance.quantity = item_godown_instance_qty + cutting_items.total_comsumption
                            item_godown_instance.save()
                            cutting_items.cutting_room_status = 'cutting_room_cancelled'
                            cutting_items.total_comsumption_in_cutting = 0
                            cutting_items.entry_from_cutting_room = True
                            cutting_items.save()

                        messages.success(request, 'Cutting Order cancelled successfully')
                        return JsonResponse({'status' : 'success'}, status=200)
                
                    else:
                        return JsonResponse({'status':'Cutting Already Approved'}, status=404)
                else:
                    return JsonResponse({'status':'Instance not found'}, status=404)
            
        except ObjectDoesNotExist as ne:
            logger.error(f'Instance not found -{ne}')
            messages.error(request, f'Error with labour workout: {ne}')
            return JsonResponse({'status':f'Instance not found -{ne}'}, status=404)
        

        except IntegrityError as ie:
            messages.error(request, 'Database integrity error occurred. Please try again.')
            logger.error(f'Database integrity error - {ie}')
            return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

        except Exception as e:
            logger.error(f'Instance not found -{e}')
            messages.error(request, f'Error with labour workout: {e}')
            return JsonResponse({'status':f'Instance not found -{e}'}, status=404)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)


@login_required(login_url='login')
def labourworkoutlistall(request):
    labour_workout_pending = labour_workout_master.objects.all().annotate(total_processed_qty = Sum('labour_workout_childs__total_process_pcs')).filter(total_pending_pcs__gt=0).order_by('created_date')
    labour_workout_completed = labour_workout_master.objects.all().annotate(total_processed_qty = Sum('labour_workout_childs__total_process_pcs')).filter(total_pending_pcs__lt=1).order_by('created_date')
    current_date = datetime.date.today
    return render(request,'production/labourworkoutlistall.html', {'labour_workout_pending':labour_workout_pending,'labour_workout_completed':labour_workout_completed,'current_date':current_date})



@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def labourworkoutsingle(request, labour_workout_child_pk=None, pk=None):

    try:
        ledger_labour_instances = Ledger.objects.filter(under_group__account_sub_group = 'Job charges(Exp of Mfg)')

        godown_id = None

        
        if pk is not None:
            
            labourworkoutinstance = get_object_or_404(labour_workout_master,id = pk)

            if labourworkoutinstance:
                godown_id = labourworkoutinstance.purchase_order_cutting_master.purchase_order_id.temp_godown_select.id

            else:
                
                raise ObjectDoesNotExist

            godown_instance = get_object_or_404(Godown_raw_material,id=godown_id)


            child_master_intial_data = {'total_approved_pcs' : labourworkoutinstance.total_approved_pcs,
                                        'total_balance_pcs' : labourworkoutinstance.total_pending_pcs}

            
            labour_work_out_child_form = labour_workout_child_form(initial=child_master_intial_data)
            
            
            product_to_item_instances = product_to_item_labour_workout.objects.filter(labour_workout = labourworkoutinstance)
        
            initial_items_data_dict = []

            for instance in product_to_item_instances:
                data_dict = {
                    'product_sku':instance.product_sku,
                    'product_color':instance.product_color,
                    'pending_pcs': instance.processed_pcs, 
                    'balance_pcs': instance.pending_pcs, 
                    'processed_pcs': 0
                    }

                initial_items_data_dict.append(data_dict)
            
            
            labour_workout_child_product_to_items_formset = inlineformset_factory(
                                    labour_workout_childs,product_to_item_labour_child_workout,fields=['product_sku',
                                                            'product_color','processed_pcs',
                                                            'pending_pcs','balance_pcs'], can_delete=False,extra=len(initial_items_data_dict))
            
            
            product_to_item_formset = labour_workout_child_product_to_items_formset(initial=initial_items_data_dict)


            
            raw_material_cutting_items_instances = purchase_order_for_raw_material_cutting_items.objects.filter(purchase_order_cutting = labourworkoutinstance.purchase_order_cutting_master).order_by('id')

            initial_data_dict = []
        
            for instance in raw_material_cutting_items_instances:

                item_instance = item_color_shade.objects.filter(items__item_name= instance.material_name,item_shade_name=instance.material_color_shade).first()

    
                if item_instance:
                    
                    current_balance = item_godown_quantity_through_table.objects.filter(godown_name = godown_instance,Item_shade_name = item_instance).first()
                    
                    if current_balance:
                        total_current_balance = current_balance.quantity
                    else:
                        total_current_balance = 0

                data = {
                    'product_sku': instance.product_sku,
                    'product_color': instance.product_color,
                    'material_name': instance.material_name,
                    'material_color_shade': instance.material_color_shade,
                    'rate': instance.rate,
                    'panha': instance.panha,
                    'units': instance.units,
                    'g_total': instance.g_total,
                    'g_total_combi' : instance.g_total_combi,
                    'consumption' : instance.consumption,
                    'combi_consumption' :instance.combi_consumption,
                    'total_comsumption': 0,
                    'unit_value': instance.unit_value,
                    'physical_stock': total_current_balance,
                    'balance_physical_stock': total_current_balance,
                    'fab_non_fab': instance.material_color_shade.items.Fabric_nonfabric,
                    'Remark': instance.Remark,
                    'pcs' : instance.pcs
                    }
                
                initial_data_dict.append(data)

            labour_workout_cutting_items_form_formset = inlineformset_factory(labour_workout_childs,labour_workout_cutting_items,
                                                                            form = labour_workout_cutting_items_form,               
                                                                            extra=len(initial_data_dict))
            
            
            labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(initial = initial_data_dict) 


        
        elif pk is None:

            labour_workout_child_instance = labour_workout_childs.objects.get(id = labour_workout_child_pk)
            

            
            labour_work_out_child_form = labour_workout_child_form(instance = labour_workout_child_instance)

            

            
            labour_workout_child_product_to_items_formset = inlineformset_factory(
                                    labour_workout_childs,product_to_item_labour_child_workout,fields=['product_sku',
                                                            'product_color','processed_pcs',
                                                            'pending_pcs','balance_pcs'], can_delete=False,extra=0)
            
            
            product_to_item_formset = labour_workout_child_product_to_items_formset(instance = labour_workout_child_instance)
            

            labour_workout_cutting_items_form_formset = inlineformset_factory(labour_workout_childs,labour_workout_cutting_items,
                                                                            form = labour_workout_cutting_items_form, 
                                                                            can_delete=False,
                                                                            extra=0)
            
            labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(instance = labour_workout_child_instance)

    except Exception as e:
        logger.error(f'exception occured at labour workout create GET {e}')



    if request.method == 'POST':
        

        
        labour_work_out_child_form = labour_workout_child_form(request.POST)

        
        product_to_item_formset = labour_workout_child_product_to_items_formset(request.POST)

        
        labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(request.POST) 


        if labour_work_out_child_form.is_valid() and product_to_item_formset.is_valid() and labour_workout_cutting_items_formset_form.is_valid():
            try:
                with transaction.atomic():
                    labour_workout_form_instance = labour_work_out_child_form.save(commit=False)
                    labour_workout_form_instance.labour_workout_master_instance = labourworkoutinstance
                    processed_qty = labour_workout_form_instance.total_process_pcs
                    labour_workout_form_instance.labour_workin_pending_pcs = processed_qty
                    labour_workout_form_instance.save()
                    
                    labourworkoutinstance.total_pending_pcs = labourworkoutinstance.total_pending_pcs - processed_qty
                
                    labourworkoutinstance.save()
                    
                    for form in product_to_item_formset:
                        if form.is_valid():
                            product_to_item_form = form.save(commit=False)
                            product_to_item_form.labour_workout = labour_workout_form_instance
                            product_to_item_form.labour_w_in_pending = product_to_item_form.processed_pcs
                            product_to_item_form.save()

                            
                            single_p2i_instance = product_to_item_instances.get(product_sku=product_to_item_form.product_sku,product_color=product_to_item_form.product_color)

                            single_p2i_instance.pending_pcs = single_p2i_instance.pending_pcs - product_to_item_form.processed_pcs
                            single_p2i_instance.save()
                    
                    
                    for form in labour_workout_cutting_items_formset_form:
                        if form.is_valid():
                            formset_form = form.save(commit=False)
                            formset_form.labour_workout_child_instance = labour_workout_form_instance
                            material_name = formset_form.material_name
                            material_color_shade = formset_form.material_color_shade

                            item_shade_instance = item_color_shade.objects.get(items__item_name=material_name,item_shade_name = material_color_shade)

                            if item_shade_instance.items.Fabric_nonfabric == 'Non Fabric':
                                total_comsumption = formset_form.total_comsumption
                                obj, created = item_godown_quantity_through_table.objects.get_or_create(godown_name=godown_instance,Item_shade_name=item_shade_instance)

                                if created:
                                    qty_to_deduct = 0
                        
                                elif not created:
                                    qty_to_deduct = obj.quantity

                                obj.quantity = qty_to_deduct - total_comsumption
                                obj.save()

                            elif item_shade_instance.items.Fabric_nonfabric == 'Fabric':
                                purchase_order_cutting_items = purchase_order_for_raw_material_cutting_items.objects.get(
                                    purchase_order_cutting__raw_material_cutting_id=labour_workout_form_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,
                                    product_sku = formset_form.product_sku, material_color_shade = item_shade_instance)
                                
                                purchase_order_cutting_items.total_comsumption_in_cutting = purchase_order_cutting_items.total_comsumption_in_cutting - formset_form.total_comsumption
                                purchase_order_cutting_items.entry_from_cutting_room = False
                                purchase_order_cutting_items.save()

                            formset_form.save()

                    return redirect(reverse('labour-workout-child-list', args=[pk]))
                
                
            except Exception as e:
                logger.error(f'An exception occoured in labour workout create {e}')

        else:
            logger.error(f'labour_workout_cutting_items_formset_form{labour_workout_cutting_items_formset_form.errors}')
            logger.error(f'product_to_item_formset {product_to_item_formset.non_form_errors()}')
            logger.error(f'labour_workout_cutting_items_formset_form - {labour_workout_cutting_items_formset_form.non_form_errors()}')

            return render(request,'production/labourworkoutsingle.html',
                  {'product_to_item_formset':product_to_item_formset,'labour_work_out_child_form':labour_work_out_child_form,
                   'labour_workout_cutting_items_formset_form':labour_workout_cutting_items_formset_form,
                   'ledger_labour_instances':ledger_labour_instances,'godown_id':godown_id})


            

    return render(request,'production/labourworkoutsingle.html',
                  {'product_to_item_formset':product_to_item_formset,'labour_work_out_child_form':labour_work_out_child_form,
                   'labour_workout_cutting_items_formset_form':labour_workout_cutting_items_formset_form,
                   'ledger_labour_instances':ledger_labour_instances,'godown_id':godown_id})




@login_required(login_url='login')
def labour_workout_child_list(request, labour_master_pk):
    labour_work_out_master = labour_workout_master.objects.get(id=labour_master_pk)
    labour_workout_child_instances = labour_workout_childs.objects.filter(labour_workout_master_instance = labour_master_pk)
    return render(request,'production/labourworkoutchilds.html', {'labour_master_pk':labour_master_pk,
                                                                  'labour_workout_child_instances':labour_workout_child_instances,
                                                                  'labour_work_out_master':labour_work_out_master})





@login_required(login_url='login')
def labourworkoutsingledeleteajax(request):
    
    if request.method == 'POST':

        labour_workout_child_pk =  request.POST.get('labour_workout_child_pk')
        try:
            with transaction.atomic():
                
                labour_workout_child_instance = labour_workout_childs.objects.get(id=labour_workout_child_pk)
                
                labour_master_instance = labour_workout_child_instance.labour_workout_master_instance
                labour_master_instance.total_pending_pcs = labour_master_instance.total_pending_pcs + labour_workout_child_instance.total_process_pcs

                labour_master_instance.save()
                
                for product_2_Item_child_instancs in labour_workout_child_instance.labour_workout_child_items.all():
                    product_2_item_master_instance = product_to_item_labour_workout.objects.get(labour_workout = labour_master_instance,product_sku=product_2_Item_child_instancs.product_sku,product_color=product_2_Item_child_instancs.product_color)
                    product_2_item_master_instance.pending_pcs = product_2_item_master_instance.pending_pcs + product_2_Item_child_instancs.processed_pcs
                    product_2_item_master_instance.save()


                for labour_workout_child_items in labour_workout_child_instance.labour_workout_cutting_items_set.all():
                    item_in_row = item_color_shade.objects.get(items__item_name=labour_workout_child_items.material_name,item_shade_name=labour_workout_child_items.material_color_shade)
                    
                    if item_in_row.items.Fabric_nonfabric == 'Fabric':
                        
                        purchase_order_cutting_item = purchase_order_for_raw_material_cutting_items.objects.get(purchase_order_cutting=labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id, material_color_shade=item_in_row, product_sku=labour_workout_child_items.product_sku)
                        
                        purchase_order_cutting_item.total_comsumption_in_cutting = purchase_order_cutting_item.total_comsumption_in_cutting + labour_workout_child_items.total_comsumption
                        purchase_order_cutting_item.entry_from_cutting_room = False
                        purchase_order_cutting_item.save()

                    elif item_in_row.items.Fabric_nonfabric == 'Non Fabric':

                        obj, created = item_godown_quantity_through_table.objects.get_or_create(Item_shade_name=item_in_row,godown_name=labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.temp_godown_select)

                        if created:
                            qty_to_add = 0
                        else:
                            qty_to_add = obj.quantity

                        obj.quantity = qty_to_add + labour_workout_child_items.total_comsumption
                        obj.save()

                labour_workout_child_instance.delete()

                messages.success(request,'labour workout deleted successfully')
                return JsonResponse({'status':'success'}, status=200) 
             

        except ObjectDoesNotExist as ne:
            messages.error(request, f'Error with labour workout: {ne}')
            logger.error(f'Instance not found - {ne}')
            return JsonResponse({'status': f'Instance not found - {ne}'}, status=404)
        

        except IntegrityError as ie:
            messages.error(request, 'Database integrity error occurred. Please try again.')
            logger.error(f'Database integrity error - {ie}')
            return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

        except Exception as e:
            logger.error(f'An unexpected error occurred - {e}')
            messages.error(request, f'Error with labour workout: {e}')
            return JsonResponse({'status': f'An unexpected error occurred - {e}'}, status=500)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)




@login_required(login_url='login')
def cuttingroomqty(request):
    cutting_room_items = purchase_order_for_raw_material_cutting_items.objects.filter(total_comsumption_in_cutting__gt=0)

    return render(request,'production/cuttingroomqty.html',{'cutting_room_items':cutting_room_items})




@login_required(login_url='login')
def labourworkincreatelist(request,l_w_o_id):

    labour_workout_child_instance = labour_workout_childs.objects.get(id=l_w_o_id)
    labour_workin_instances = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).annotate(approved_Qty_total=Sum('l_w_in_products__approved_qty'),total_approved_pcs = Sum('l_w_in_products__approved_qty'),pending_for_approval_pcs = Sum('l_w_in_products__pending_for_approval'))
    return render(request,'production/labour_work_in_list.html',{'labour_workout_child_instance':labour_workout_child_instance,
                                                                 'labour_workin_instances':labour_workin_instances})




@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def labourworkincreate(request, l_w_o_id = None, pk = None, approved=False):
    
    template_name = 'production/labourworkincreate.html'
    

    approval_check = approved
    
    
    if l_w_o_id is None:

        template_name = 'production/labourworkincreateraw.html'
        labour_workin_master_instance = None
        
        master_form = labour_workin_master_form()

        product_to_item_formset = None

        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item, 
            form = labour_work_in_product_to_item_form, extra = 0, can_delete = False)

        labour_workout_child_instance = None

        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            
            try:
                vendor_name_value = request.GET.get('nameValue')

                vendor_name_dict = None
    
                if vendor_name_value:
                    selected_vendor_name = Ledger.objects.filter(under_group__account_sub_group='Job charges(Exp of Mfg)',name__icontains=vendor_name_value)
                    vendor_name_dict = {}
                    for record in selected_vendor_name:
                        vendor_name_dict[record.id] = record.name

                choosed_vendor_name = request.GET.get('itemValue')       


                labour_workout_instance_dict = []

                if choosed_vendor_name:
                    labour_workout_instances = labour_workout_childs.objects.filter(labour_name=choosed_vendor_name)
                    
                    for instance in labour_workout_instances:
                        dict_to_append = {
                            'Challan_No': instance.challan_no,
                            'PO_No':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                            'PO_Total_QTY':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,
                            'Ref_No':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,
                            'Model_Name':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,
                            'Issued_QTY':instance.total_process_pcs,
                            'Rec_QTY': instance.labour_workin_pcs,
                            'Balance_QTY': instance.labour_workin_pending_pcs,
                            'labour_workout_id': instance.id}

                        labour_workout_instance_dict.append(dict_to_append)

                
                labour_work_out_id = request.GET.get('labourWorkOutId')
               
                master_initial_data = None

                formset_initial_data = None

                labour_workout_child_instance_id = None

                if labour_work_out_id:
                    labour_workout_child_instance = labour_workout_childs.objects.get(id = labour_work_out_id)
                    labour_workout_child_instance_id = labour_workout_child_instance.id

                    master_initial_data = {
                        'labour_name': labour_workout_child_instance.labour_name.name,
                        'challan_no' : labour_workout_child_instance.challan_no ,
                        'purchase_order_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                        'refrence_number' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,
                        'model_name': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,
                        'total_p_o_qty' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,
                        'labour_workout_qty' : labour_workout_child_instance.total_process_pcs,
                        'labour_charges': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.labour_charges,
                        'total_balance_pcs' :  labour_workout_child_instance.labour_workin_pending_pcs
                        }

                    product_to_item_l_w_in_instance = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)

                    
                    formset_initial_data = []

                    for instances in product_to_item_l_w_in_instance:

                        initial_data_dict = { 
                            'product_sku': instances.product_sku,
                            'product_color': instances.product_color,
                            'L_work_out_pcs': instances.processed_pcs,
                            'pending_to_return_pcs': instances.labour_w_in_pending,
                            'return_pcs' : '0',
                            'qty_to_compare':  instances.labour_w_in_pending,
                            'cur_bal_plus_return_qty': instances.labour_w_in_pending 
                        }

                        formset_initial_data.append(initial_data_dict)
                
                return JsonResponse({'vendor_name_dict':vendor_name_dict,'labour_workout_instance_dict':labour_workout_instance_dict,
                                     'master_initial_data':master_initial_data,'formset_initial_data':formset_initial_data,
                                     'labour_workout_child_instance_id': labour_workout_child_instance_id})

            except ValueError as ve:
                    messages.error(request,f'Error Occured - {ve}')
                    return JsonResponse({'status': f'Error with ajax request - {ve}'}, status=404)
        
            except Exception as e:
                messages.error(request,f'Exception Occured - {e}')
                return JsonResponse({'status': f'Error with ajax request - {e}'}, status=404)


    
    elif l_w_o_id is not None and pk is None:

        labour_workin_master_instance = None
        labour_workout_child_instance = labour_workout_childs.objects.get(id=l_w_o_id)

        initial_data = {
            'labour_name': labour_workout_child_instance.labour_name.name,
            'challan_no' : labour_workout_child_instance.challan_no ,
            'purchase_order_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
            'refrence_number' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,
            'model_name': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,
            'total_p_o_qty' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,
            'labour_workout_qty' : labour_workout_child_instance.total_process_pcs,
            'labour_charges': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.labour_charges,
            'total_balance_pcs' :  labour_workout_child_instance.labour_workin_pending_pcs,  
        }

        master_form = labour_workin_master_form(initial=initial_data)

        product_to_item_l_w_in = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)

        formset_initial_data = []

        for instances in product_to_item_l_w_in:

            initial_data_dict = { 
                'product_sku': instances.product_sku,
                'product_color': instances.product_color,
                'L_work_out_pcs': instances.processed_pcs,
                'pending_to_return_pcs': instances.labour_w_in_pending,
                'return_pcs' : '0',
                'qty_to_compare':  instances.labour_w_in_pending,
                'cur_bal_plus_return_qty': instances.labour_w_in_pending
                }
            
            formset_initial_data.append(initial_data_dict)


        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item, 
            form=labour_work_in_product_to_item_form, extra=len(formset_initial_data), can_delete=False)

        product_to_item_formset = labour_work_in_product_to_item_formset(initial=formset_initial_data)


    
    elif l_w_o_id is not None and pk is not None:
    
        labour_workout_child_instance = labour_workout_childs.objects.get(id = l_w_o_id)

        product_to_item_l_w_in = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)

        labour_workin_master_instance = labour_work_in_master.objects.get(pk=pk)
    
        master_form = labour_workin_master_form(instance = labour_workin_master_instance)

        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item, 
            form = labour_work_in_product_to_item_form, extra = 0, can_delete = False)
        
        product_to_item_formset = labour_work_in_product_to_item_formset(instance = labour_workin_master_instance) 
        
        
        for form, instance in zip(product_to_item_formset.forms, product_to_item_l_w_in):  

            if instance:
                form.initial['qty_to_compare'] = instance.labour_w_in_pending
                form.initial['cur_bal_plus_return_qty'] = instance.labour_w_in_pending  + form.instance.return_pcs

    if request.method == 'POST':
        labour_workout_child_i = request.POST.get('labour_workout_child_instance_id')
        if labour_workout_child_i:
            labour_workout_child_instance = labour_workout_childs.objects.get(id=labour_workout_child_i)

        master_form = labour_workin_master_form(request.POST, instance = labour_workin_master_instance)
        product_to_item_formset = labour_work_in_product_to_item_formset(request.POST,instance = labour_workin_master_instance)

        try:
            with transaction.atomic():
                if master_form.is_valid() and product_to_item_formset.is_valid():
                    parent_form = master_form.save(commit = False)
                    parent_form.labour_voucher_number = labour_workout_child_instance

                    
                    labour_workout_child_instance.labour_workin_pcs = labour_workout_child_instance.labour_workin_pcs + parent_form.total_return_pcs

                    parent_form.labour_voucher_number.labour_workin_pending_pcs = parent_form.total_balance_pcs

                    labour_workout_child_instance.save()

                    parent_form.save()

                    for form in product_to_item_formset:

                        if form.is_valid():
                            product_to_item_form = form.save(commit= False)
                            product_to_item_form.labour_workin_instance = parent_form

                            product_to_item_form.pending_for_approval = product_to_item_form.return_pcs
                            
                            l_w_o_instance = product_to_item_labour_child_workout.objects.get(labour_workout=labour_workout_child_instance,
                                                                                              product_sku=product_to_item_form.product_sku,
                                                                                              product_color=product_to_item_form.product_color)
                            
                            
                            if product_to_item_form.pk:

                                labour_workin_product2item = labour_work_in_product_to_item.objects.get( pk = product_to_item_form.pk)
                                qty_to_change = product_to_item_form.return_pcs - labour_workin_product2item.return_pcs

                            else:
                                qty_to_change = product_to_item_form.return_pcs

                            l_w_o_instance.labour_w_in_pending = l_w_o_instance.labour_w_in_pending - qty_to_change

                            l_w_o_instance.save()
                            product_to_item_form.save()
                    
                    return redirect(reverse('labour-workin-list-create', args=[labour_workout_child_instance.id]) )

                else:
                    
                    
                    
                    
                    return redirect(reverse('labour-workin-list-create', args=[labour_workout_child_instance.id]) )
                    
                
                
        except ValidationError as ve:
            messages.error(request,f'Validation error {ve}')

        except Exception as e:
            messages.error(request,f'Other exceptions {e}')

    return render(request,template_name,{'master_form':master_form,'labour_work_in_product_to_item_formset':product_to_item_formset,'approval_check':approval_check})


@login_required(login_url='login')
def labourworkinlistall(request):

    
    labour_workout_childs_exists = labour_workout_childs.objects.filter(
    labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id = OuterRef('pk')
    ).values('pk')[:1]

    
    purchase_orders_with_labour_workout_childs = purchase_order.objects.annotate(
    has_labour_workout_childs=Exists(labour_workout_childs_exists)
    ).filter(has_labour_workout_childs = True).annotate(total_lwo_pcs = Sum('cutting_pos__labourworkouts__labour_workout_childs__total_process_pcs'),
    total_labour_workin_pcs = Sum('cutting_pos__labourworkouts__labour_workout_childs__labour_workin_pcs'),
    total_labour_workin_pending = Sum('cutting_pos__labourworkouts__labour_workout_childs__labour_workin_pending_pcs'))


    labour_workout_child_instances_all = labour_workout_childs.objects.all()

    return render(request,'production/labour_workin_listall.html',
                  {'labour_workout_child_instances_all':labour_workout_child_instances_all,
                   'purchase_order_instances': purchase_orders_with_labour_workout_childs})




@login_required(login_url='login')
def labourworkinpurchaseorderlist(request,p_o_no):
    purchase_order_instance = purchase_order.objects.get(id=p_o_no)

    labour_workin_purchase_order_list = labour_workout_childs.objects.filter(labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__id = p_o_no)

    return render(request,'production/labour_workin_purchase_order_list.html',{'labour_workin_purchase_order_list':labour_workin_purchase_order_list,'purchase_order_instance':purchase_order_instance})




@login_required(login_url='login')
def labourworkinsingledeleteajax(request):
    
    if request.method == 'POST':
        labour_workin_id = request.POST.get('labour_workin_pk')

        if labour_workin_id:
            try:
                with transaction.atomic():
                    labour_workin_instance = labour_work_in_master.objects.get(pk=labour_workin_id)
                    
                    labour_workin_instance.labour_voucher_number.labour_workin_pending_pcs = labour_workin_instance.labour_voucher_number.labour_workin_pending_pcs + labour_workin_instance.total_return_pcs
                    labour_workin_instance.labour_voucher_number.labour_workin_pcs = labour_workin_instance.labour_voucher_number.labour_workin_pcs - labour_workin_instance.total_return_pcs
                    labour_workin_instance.labour_voucher_number.save()


                    for instances in labour_workin_instance.l_w_in_products.all():
                        product_2_item_child_instances =  product_to_item_labour_child_workout.objects.get(labour_workout=labour_workin_instance.labour_voucher_number,product_sku=instances.product_sku,product_color=instances.product_color)
                        product_2_item_child_instances.labour_w_in_pending = product_2_item_child_instances.labour_w_in_pending + instances.return_pcs

                        product_2_item_child_instances.save()

                    labour_workin_instance.delete()

                    return redirect(reverse('labour-workin-list-create', args=[labour_workin_instance.labour_voucher_number.id]) )
                    

            except ObjectDoesNotExist as ne:
                messages.error(request, f'Error with labour workout: {ne}')
                logger.error(f'Instance not found - {ne}')
                return JsonResponse({'status': f'Instance not found - {ne}'}, status=404)
        

            except IntegrityError as ie:
                messages.error(request, 'Database integrity error occurred. Please try again.')
                logger.error(f'Database integrity error - {ie}')
                return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

            except Exception as e:
                logger.error(f'An unexpected error occurred - {e}')
                messages.error(request, f'Error with labour workout: {e}')
                return JsonResponse({'status': f'An unexpected error occurred - {e}'}, status=500)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)


@login_required(login_url='login')
def goods_return_pending_list(request):

    labour_workin_instances = labour_work_in_master.objects.all().annotate(total_approved_pcs = Sum('l_w_in_products__approved_qty'),pending_for_approval_pcs = Sum('l_w_in_products__pending_for_approval'))
    return render(request,'production/goodsreturnpendinglist.html',{'labour_workin_instances':labour_workin_instances})




@login_required(login_url='login')
def goods_return_popup(request,pk):

    if pk:
        finished_goods_godowns = Godown_finished_goods.objects.all()
        labour_workin_instance = labour_work_in_master.objects.get(pk=pk)
        formset = labour_work_in_product_to_item_approval_formset(request.POST or None, instance=labour_workin_instance)

        if request.method == 'POST':
            godown_name_post = request.POST.get('godown_name_post')
            formset.forms = [form for form in formset if form.has_changed()]

            if formset.is_valid():
                with transaction.atomic():
                    try:
                        for form in formset:
                            form_instance = form.save(commit=False)
            
                            godown_instance =  Godown_finished_goods.objects.get(id=godown_name_post)
                            selected_product = PProduct_Creation.objects.get(PProduct_SKU = form_instance.product_sku)
                            obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance, product_color_name = selected_product)

                            if created:
                                quantity_to_add = 0

                            else:
                                quantity_to_add = obj.quantity
                            
                            if form_instance.pk:
                                labour_workin_instance = labour_work_in_product_to_item.objects.get(pk = form_instance.pk)

                                approved_qty_differ = form_instance.approved_qty - labour_workin_instance.approved_qty 
                            else:
                                approved_qty_differ = form_instance.approved_qty

                            obj.quantity = quantity_to_add + approved_qty_differ  

                            obj.save()
                            form_instance.save()

                        messages.success(request,'Products Successfully recieved')

                    except Exception as e:
                        messages.error(request,f'Error with Formset - {e}')


                
                close_window_script = """
                <script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>
                """
                return HttpResponse(close_window_script)

            else:
                messages.error(request,f'Error with Formset {formset.errors}')

                
                close_window_script = """
                <script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>
                """
                return HttpResponse(close_window_script)

    return render(request,'production/goods_return_popup.html',{'formset':formset,'finished_goods_godowns':finished_goods_godowns})



def finished_goods_godown_wise_report(request, g_id):
    
    
    all_godown_stock = product_godown_quantity_through_table.objects.filter(
        product_color_name__Product_id = OuterRef('pk')).values('product_color_name__Product_id'
        ).annotate(sum_all_g_qty=Sum('quantity')).values('sum_all_g_qty')

    
    product_quantity = Product.objects.filter(
        productdetails__godown_colors__godown_name__id=g_id).annotate(
        total_quantity_product=Sum('productdetails__godown_colors__quantity'),  
        all_godown_qty = Subquery(all_godown_stock))  

    return render(request,'production/godown_product_qty.html', {'product_quantity' : product_quantity})


def finished_goods_godown_product_ref_wise_report(request, ref_no):

    if ref_no:
        product_instance = Product.objects.get(Product_Refrence_ID = ref_no)

        purchase_instances = labour_work_in_master.objects.filter(
            labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no).annotate(
            total_balance_to_vendor = F('labour_voucher_number__total_process_pcs') - F('total_return_pcs')    
            )

    return render(request,'production/godown_model_wise.html', {'purchase_instances' : purchase_instances,'product_instance':product_instance})


def finished_goods_vendor_model_wise_report(request, vendor_id, ref_no):


    if ref_no is not None and vendor_id is not None:
        queryset_list = []

        labour_workout_instances = product_to_item_labour_child_workout.objects.filter(labour_workout__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no, labour_workout__labour_name__id=vendor_id)

        labour_workin_instances = labour_work_in_product_to_item.objects.filter(labour_workin_instance__labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no,labour_workin_instance__labour_voucher_number__labour_name__id=vendor_id)

        product_instance = get_object_or_404(Product,Product_Refrence_ID=ref_no)
        reference_no = ref_no
        model_number = product_instance.Model_Name

        for instance in labour_workout_instances:
            dict_to_append = {
                'GRN_No':instance.labour_workout.challan_no,
                'date':instance.labour_workout.created_date,
                'description' : 'LWI',
                'SKU_No' : instance.product_sku,
                'Quantity': instance.processed_pcs
            }
            queryset_list.append(dict_to_append)

        for instance in labour_workin_instances:

            dict_to_append = {
                'GRN_No' : instance.labour_workin_instance.voucher_number,
                'date':instance.labour_workin_instance.created_date,
                'description': 'LWO',
                'SKU_No': instance.product_sku,
                'Quantity': instance.return_pcs
            }
            queryset_list.append(dict_to_append)

        report_data_sorted = sorted(queryset_list, key = itemgetter('date'), reverse=False)

        return render(request,'production/finishedgoodsvendormodelwisereport.html',{'report_data_sorted':report_data_sorted, 'reference_no':reference_no, 'model_number':model_number})







@login_required(login_url='login')
def factory_employee_create_update_list(request ,pk=None):
    

    if request.user.is_superuser:
        factory_employees = factory_employee.objects.all()
        cutting_rooms =  cutting_room.objects.all()
    
        
        



    if pk:
        title = 'Update'
        instance = get_object_or_404(factory_employee,pk=pk)

    else:
        title = 'Create'
        instance = None
    
    form = factory_employee_form(request.POST or None, instance = instance)

    if request.method == 'POST':
        if form.is_valid():
            form.save()

            messages.success(request,'Factory Employee created Successfully')
            return redirect('factory-emp-create')
        else:
            messages.error(request,f'Error with form submission {form.errors}')

    return render(request,'production/factory_emp_create_update_list.html', {'form':form,'factory_employees':factory_employees,'title':title,'cutting_rooms':cutting_rooms})



@login_required(login_url='login')
def factoryempdelete(request,pk=None): 
    try:
        instance = get_object_or_404(factory_employee,pk=pk)
        instance.delete()
        messages.success(request,f'Factory Employee {instance.factory_emp_name} was deleted')

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {instance.factory_emp_name} because it is referenced by other objects.')
    return redirect('factory-emp-create')


@login_required(login_url='login')
def cutting_room_create_update_list(request, pk=None):
    print(request.POST)
    if pk:
        instance = cutting_room.objects.get(id = pk)

    else:
        instance = None

    form = cutting_room_form(request.POST or None, instance = instance) 

    if request.user.is_superuser:
        cutting_rooms = cutting_room.objects.all()
    
    


    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('cutting_room-create')
    
    return render(request,'production/cuttingroomcreateupdatelist.html', {'form':form,'cutting_rooms':cutting_rooms})



@login_required(login_url='login')
def cuttingroomdelete(request,pk):
    instance = cutting_room.objects.get(pk=pk)
    instance.delete()
    return redirect('cutting_room-create')







@login_required(login_url='login')
def itemdynamicsearchajax(request):
    
    try:
        
        item_name_typed = request.GET.get('nameValue')

        if not item_name_typed:
            raise ValidationError("partial name provided.")
        
        logger.info(f"searched keyword via itemdynamicsearchajax {item_name_typed}")
        
        item_name_searched = Item_Creation.objects.filter(item_name__icontains=item_name_typed).annotate(total_qty= Sum('shades__godown_shades__quantity'))
        

        if item_name_searched:
            
            

            

            searched_item_name_dict = {}
            for queryset in item_name_searched:

                if queryset.total_qty is not None:
                    total_qty = str(queryset.total_qty)
                else:
                    total_qty = '0'

                item_name = queryset.item_name + ' | ' + total_qty
                item_id = queryset.id
                searched_item_name_dict[item_id] = item_name
            
            logger.info(f"searched result via itemdynamicsearchajax {searched_item_name_dict}")
            
            return JsonResponse({'item_name_typed': item_name_typed, 'searched_item_name_dict': searched_item_name_dict}, status=200)
        else:
            return JsonResponse({'error': 'No items found.'}, status=404)

    except ValidationError as ve:
        error_message = str(ve)
        logger.error(f"Validaton errorin itemdynamicsearchajax - {ve}")
        return JsonResponse({'error': error_message}, status=400)
    
    except Exception as e:
        logger.error(f"Exception in itemdynamicsearchajax - {ve}")
        error_message = f"An error occurred:{str(e)}"
        return JsonResponse({'error': error_message}, status=500)



def CheckUniqueFieldDuplicate(model_name, searched_value, col_name):
    
    if searched_value:
        validation_flag = False
        try:
            
            lookup = {f"{col_name}__iexact": searched_value}
            check_instance_valid = model_name.objects.get(**lookup)
            
            validation_flag = True

        except model_name.DoesNotExist:
            validation_flag = False
            
        except Exception as e:
            return JsonResponse({f'Status':'Exception Occoured - {e}'}, status=404)
        
        return JsonResponse({'validation_flag':validation_flag})
    else:
        return JsonResponse({f'Status':'No data recieved - {e}'}, status=404)


@login_required(login_url='login')
def UniqueValidCheckAjax(request):
    searched_from = request.GET.keys()
    
    if 'purchase_number' in searched_from:
        searched_value = request.GET.get('purchase_number').strip()
        print(searched_value)
        model_name = item_purchase_voucher_master
        col_name = 'purchase_number'

    if 'new_order_number' in searched_from:
        searched_value = request.GET.get('new_order_number').strip()
        model_name = purchase_order
        col_name = 'purchase_order_number'

    elif 'cutting_order_number' in searched_from:
        searched_value = request.GET.get('cutting_order_number').strip()
        model_name = purchase_order_raw_material_cutting
        col_name = 'raw_material_cutting_id'

    elif 'labour_workout_challan_no' in searched_from:
        searched_value = request.GET.get('labour_workout_challan_no').strip()
        model_name = labour_workout_childs
        col_name = 'challan_no'

    elif 'item_name' in searched_from:
        searched_value = request.GET.get('item_name').strip()
        model_name = Item_Creation
        col_name = 'item_name'

    elif 'item_material_code' in searched_from:
        searched_value = request.GET.get('item_material_code').strip()
        model_name = Item_Creation
        col_name = 'Material_code'
    
    elif 'voucher_number' in searched_from:
        searched_value = request.GET.get('voucher_number').strip()
        model_name = labour_work_in_master
        col_name = 'voucher_number'

    else:
        model_name = None
        searched_value = None
        col_name = None


    
    if model_name and searched_value and col_name:
        return CheckUniqueFieldDuplicate(model_name, searched_value, col_name)
    else:
        return JsonResponse({'Status': 'Invalid data received'}, status=400)


            
@login_required(login_url='login')
def session_data_test(request):
    
    
    
    

    
    session_data = request.session
    
    
    
    for key, value in session_data.items():
        print(f"Key: {key}, Value: {value}")

    context = {}
    return render(request,'misc/session_test.html',context=context)




@login_required(login_url='login')
def creditdebitreport(request):
    all_reports = account_credit_debit_master_table.objects.all()

    return render(request,'misc/credit_debit_master_report.html',{'all_reports':all_reports})



@login_required(login_url='login')
def godown_stock_raw_material_report_fab_grp(request,g_id,fab_id=None):
    
    
    items_in_godown = item_godown_quantity_through_table.objects.filter(godown_name=g_id)
    
    
    Fabric_grp_name = None
    querylist = None

    
    if fab_id:
        page_id = 'item_page'

    else:
        page_id = 'fabric_page'
    
    if page_id == 'fabric_page':
        
        fabric_in_godown = items_in_godown.distinct('Item_shade_name__items__Fabric_Group')
        
        list_fab_grp = []

        
        for fab in fabric_in_godown:
            list_fab_grp.append(fab.Item_shade_name.items.Fabric_Group.id)

        queryset = []

        
        for items in list_fab_grp:
            values = Fabric_Group_Model.objects.filter(id=
                    items).filter(items__shades__godown_shades__godown_name=g_id).annotate(total_qty = 
                    Round(Sum('items__shades__godown_shades__quantity'), 2),
                    avg_rate=Round(Avg('items__shades__rate'),2)).first()

                    

            queryset.append(values)

    elif page_id == 'item_page':

        

        Fabric_grp_name = Fabric_Group_Model.objects.get(id=fab_id)

        items_in_fab_grp = Item_Creation.objects.filter(Fabric_Group=fab_id).filter(
            shades__godown_shades__godown_name__id=g_id).annotate(
                total_qty =Round(Sum('shades__godown_shades__quantity')))
                

        querylist = []
        
        
        for query in items_in_fab_grp: 
            item_dict = {}
            item_dict['item_name'] = query.item_name
            item_dict['total_qty'] = query.total_qty

            shades_list = []
            for shade in query.shades.filter(godown_shades__godown_name__id=g_id): 
                shade_dict = {}
                shade_dict['rate'] = shade.rate
                shades_list.append(shade_dict)

                shade_godown_list = []
                for godown_items in shade.godown_shades.filter(godown_name__id=g_id): 
                    godown_shade_dict = {}
                    shade_dict['item_shade'] = godown_items.Item_shade_name.item_shade_name
                    shade_dict['item_shade_id'] = godown_items.Item_shade_name.id
                    shade_dict['quantity'] = godown_items.quantity
                    shade_godown_list.append(godown_shade_dict)

            item_dict['shades'] = shades_list
            querylist.append(item_dict)
        queryset = items_in_fab_grp

    godown_name = Godown_raw_material.objects.get(id = g_id)
    
    
    return render(request,'reports/godownstockrawmaterialreportfabgrp.html',{'page_id':page_id,
                                                                             'godown_id':g_id,
                                                                             'godown_name':godown_name,
                                                                             'Fabric_grp_name':Fabric_grp_name,
                                                                             'queryset':queryset,
                                                                             'querylist':querylist})



@login_required(login_url='login')
def godown_item_report(request,shade_id,g_id=None):
    
    shade_name = item_color_shade.objects.get(id=shade_id)
    godown_name = 'All Stock'

    report_data = []

    if g_id is not None:
        godown_name = Godown_raw_material.objects.get(id=g_id)


        
        opening_godown_qty = opening_shade_godown_quantity.objects.filter(
            opening_purchase_voucher_godown_item=shade_name, opening_godown_id=godown_name)
        

        
        
        purchase_voucher_godown_qty = item_purchase_voucher_master.objects.filter(
            purchase_voucher_items__item_shade = shade_name , purchase_voucher_items__shade_godown_items__godown_id = godown_name).annotate(
                godown_qty_total=Sum('purchase_voucher_items__shade_godown_items__quantity'), item_rate=Round(Avg(
                    'purchase_voucher_items__rate')), filter=Q(purchase_voucher_items__shade_godown_items__godown_id = godown_name))
        

        
        purchase_order_cutting_room_qty = godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = False, godown_id = g_id)
        
        
    
        
        purchase_order_cutting_room_qty_cancelled =  godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = True, godown_id = g_id)
        

        labour_workout_report = labour_workout_cutting_items.objects.filter(material_name = shade_name.items.item_name,
            material_color_shade = shade_name.item_shade_name,labour_workout_child_instance__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__temp_godown_select = g_id)
        

    else:
    
        
        opening_godown_qty = opening_shade_godown_quantity.objects.filter(
            opening_purchase_voucher_godown_item=shade_name)
        


        
        
        purchase_voucher_godown_qty = item_purchase_voucher_master.objects.filter(
            purchase_voucher_items__item_shade = shade_name).annotate(
                godown_qty_total=Sum('purchase_voucher_items__shade_godown_items__quantity'), item_rate=Round(Avg(
                    'purchase_voucher_items__rate')))
        

        
        purchase_order_cutting_room_qty = godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = False)
        
        
    
        
        purchase_order_cutting_room_qty_cancelled =  godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = True)
        

        labour_workout_report = labour_workout_cutting_items.objects.filter(material_name = shade_name.items.item_name,
            material_color_shade = shade_name.item_shade_name)
    
        
    for godown_qty in opening_godown_qty:

        report_data.append({
            'date': godown_qty.created_date,
            'particular': 'Opening Balance',
            'voucher_type': '',
            'vch_no': '',
            'inward_quantity': f"{godown_qty.opening_quantity}",
            'inward_value': godown_qty.opening_rate * godown_qty.opening_quantity,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity': 0,
            'closing_value': 0,
            'rate':godown_qty.opening_rate
        })

    for purchase_voucher_item_qty in purchase_voucher_godown_qty:

        report_data.append({
            'date': purchase_voucher_item_qty.created_date,
            'particular': 'Puchase Voucher',
            'voucher_type': purchase_voucher_item_qty.ledger_type,
            'vch_no': purchase_voucher_item_qty.purchase_number,
            'inward_quantity': f"{purchase_voucher_item_qty.godown_qty_total}",
            'inward_value': purchase_voucher_item_qty.item_rate * purchase_voucher_item_qty.godown_qty_total,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity':0,
            'closing_value': 0,
            'rate': purchase_voucher_item_qty.item_rate
            })
        

    for fabric_cutting_items in purchase_order_cutting_room_qty:
        outward_value = round(fabric_cutting_items.total_comsumption * fabric_cutting_items.rate , 2)
        report_data.append({
            'date': fabric_cutting_items.creation_date,
            'particular': 'Cutting Room',
            'voucher_type': 'Cutting Room',
            'vch_no': fabric_cutting_items.voucher_number,
            'inward_quantity': '',
            'inward_value': '',
            'outward_quantity': f"{fabric_cutting_items.total_comsumption}",
            'outward_value': outward_value,
            'closing_quantity': 0,
            'closing_value': 0,
            'rate': fabric_cutting_items.rate})
    

    for fabric_cutting_cancelled_items in purchase_order_cutting_room_qty_cancelled:
        inward_value = round(fabric_cutting_cancelled_items.total_comsumption * fabric_cutting_items.rate, 2)
        report_data.append({
            'date': fabric_cutting_cancelled_items.creation_date,
            'particular': 'Cutting Room Cancelled',
            'voucher_type': 'Cutting Room',
            'vch_no': fabric_cutting_cancelled_items.voucher_number,
            'inward_quantity': f"{fabric_cutting_cancelled_items.total_comsumption}",
            'inward_value': inward_value,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity': 0,
            'closing_value': 0,
            'rate': fabric_cutting_cancelled_items.rate})
        
    for record in labour_workout_report:
        item_instance = item_color_shade.objects.get(items__item_name=record.material_name,item_shade_name=record.material_color_shade)
        
        if item_instance.items.Fabric_nonfabric == 'Non Fabric':
            outward_value = round(record.total_comsumption * record.rate , 2)

            report_data.append({
                'date': record.created_date,
                'particular': 'Labour workout',
                'voucher_type': 'Labour workout',
                'vch_no': record.labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                'inward_quantity': '',
                'inward_value': '',
                'outward_quantity': f"{record.total_comsumption}",
                'outward_value': outward_value,
                'closing_quantity': 0,
                'closing_value': 0,
                'rate': record.rate})
    
    report_data_sorted = sorted(report_data, key = itemgetter('date'), reverse=False)

    """
    The sorted() function is used to sort the list. It returns a new list that is sorted 
    according to the specified key.

    The key parameter specifies a function that is used to extract the sorting key from each
    dictionary. Here, itemgetter('date') is used to get the value associated with the date
    key in each dictionary.

    If you set reverse=False, it would sort the list in ascending order.

    """
    
    return render(request,'reports/godownstockrawmaterialreportsingle.html',{'godoown_name':godown_name,
                                                                             'shade_name':shade_name,'report_data':report_data_sorted})




@login_required(login_url='login')
def allrawmaterialstockreport(request):
    queryset = item_color_shade.objects.all().annotate(total_qty = Sum(
        'godown_shades__quantity')).order_by('items__item_name').prefetch_related(
            'godown_shades','godown_shades__godown_name').select_related('items')
    
    return render(request,'reports/allrawmaterialstockreport.html',{'queryset':queryset})




@login_required(login_url='login')
def allfinishedgoodsstockreport(request):
    product_queryset = PProduct_Creation.objects.all().annotate(total_qty = Sum(
        'godown_colors__quantity')).order_by('Product__Model_Name')
    
    return render(request,'reports/allfinishedgoodsstockreport.html',{'product_queryset':product_queryset})


@login_required(login_url='login')
def raw_material_excel_download(request):

    wb = Workbook()

    default_sheet = wb['Sheet']
    wb.remove(default_sheet) 

    wb.create_sheet('raw_material_create')
    sheet1 = wb.worksheets[0]
    headers =  ['Raw Material Name', 'Material Code','Color','Packing','Unit Name','Units','Panha', 
                'Fabric or Non Fabric','Fabric Finishes','Fabric Group','GST','HSN Code','Status']
    
    sheet1.append(headers)

    fileoutput = BytesIO()
    wb.save(fileoutput)
        
    
    response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = 'raw_material_create'
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    return response



@login_required(login_url='login')
def raw_material_excel_upload(request):
    
    if request.method == "POST":
        try:
            excel_file = request.FILES.get('excel_file')
        
            if excel_file:
                file_name = excel_file.name.split('.')[0].split(' ')[0]

            
                if file_name == 'raw_material_create' and excel_file.name.endswith('.xlsx'):
                    
                    df = pd.read_excel(excel_file)
                    
                    
                    required_columns = {
                        'Raw Material Name':'item_name',
                        'Material Code':'Material_code',
                        'Color':'Item_Color',
                        'Packing':'Item_Packing',
                        'Unit Name':'unit_name_item',
                        'Panha':'Panha', 
                        'Units': 'Units',
                        'Fabric or Non Fabric':'Fabric_nonfabric',
                        'Fabric Finishes':'Item_Fabric_Finishes',
                        'Fabric Group':'Fabric_Group',
                        'GST':'Item_Creation_GST',
                        'HSN Code':'HSN_Code',
                        'Status':'status'
                    }
                    
                    
                    for col in required_columns.keys():
                        if col not in df.columns:
                            return HttpResponse(f"Missing required column: {col}")
                        

                    rows_with_error = []
                    for index, row in df.iterrows():
                        with transaction.atomic():
                            try:
                                
                                color = Color.objects.get(color_name=row['Color'])
                                packaging_m = packaging.objects.get(packing_material=row['Packing'])
                                unit_name = Unit_Name_Create.objects.get(unit_name=row['Unit Name'])
                                fabric_finish = FabricFinishes.objects.get(fabric_finish=row['Fabric Finishes'])
                                fabric_group = Fabric_Group_Model.objects.get(fab_grp_name=row['Fabric Group'])
                                gst_instance = gst.objects.get(gst_percentage=row['GST'])
                                
                                material_name = Item_Creation.objects.filter(item_name=row['Raw Material Name']).first()
                                material_code = Item_Creation.objects.filter(Material_code=row['Material Code']).first()

                                if not material_name and not material_code:
                                    
                                    Item_Creation.objects.create(item_name=row['Raw Material Name'],
                                                                            Material_code= row['Material Code'],
                                                                            Item_Color = color,
                                                                            Item_Packing = packaging_m,
                                                                            unit_name_item = unit_name,
                                                                            Units =  row['Units'],
                                                                            Panha =  row['Panha'],
                                                                            Fabric_nonfabric = row['Fabric or Non Fabric'],
                                                                            Item_Fabric_Finishes = fabric_finish,
                                                                            Fabric_Group = fabric_group,
                                                                            Item_Creation_GST = gst_instance,
                                                                            HSN_Code = row['HSN Code'],
                                                                            status = row['Status']
                                                                            )
                                else:
                                    logger.error(f'Duplicate  material_name{material_name} or material_code {material_code}')
                                    raise ValidationError(f'Duplicate  material_name{material_name} or material_code {material_code}')


                            except ValidationError as ve:
                                logger.error(f'validation error {ve}')
                                rows_with_error.append(row)

                            except Exception as e:
                                rows_with_error.append(row)
                                messages.error(request, f"Error processing row {index + 1}: {str(e)}")
                                logger.error(f"Error processing row {index + 1}: {str(e)}")
                                

                    if rows_with_error:

                        
                        list_of_lists = [df.values.tolist() for df in rows_with_error]


                        wb = Workbook()
                        default_sheet = wb['Sheet']
                        wb.remove(default_sheet) 

                        wb.create_sheet('raw_material_create_with_errors')

                        sheet1 = wb.worksheets[0]
                        headers =  ['Raw Material Name', 'Material Code','Color', 'Packing','Unit Name','Units','Panha', 
                                'Fabric or Non Fabric','Fabric Finishes','Fabric Group','GST','HSN Code','Status']
    
                        sheet1.append(headers)

                        for row in list_of_lists:
                            sheet1.append(row)


                        fileoutput = BytesIO()
                        wb.save(fileoutput)
        
                        
                        response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        file_name = 'raw_material_create_with_errors'
                        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
                        logging.error(f"Item saved successfully with errors no of errors {len(rows_with_error)}")
                        messages.error(request, f"Item saved successfully with errors no of errors {len(rows_with_error)}")
                        return response
                            
                    messages.success(request, "Item saved successfully")
                    return redirect('item-list')
            
                else:
                    messages.error(request, 'Invalid file format. Please upload a valid Excel file.')
                    logger.error('Invalid file format-item excel. Please upload a valid Excel file.')
                    raise ValidationError('Invalid file format-item excel. Please upload a valid Excel file.')
        
        except Exception as e:
            messages.error(request, f'Invalid file format. Please upload a valid Excel file.{e}')
            logger.error(f'Invalid file format-item excel. Please upload a valid Excel file.{e}')
            return redirect('item-list')
    else:
        messages.error(request, "Invalid request method")
        











